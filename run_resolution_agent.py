"""
Resolution Agent — Downloads afternoon/evening pending report and
compares ticket IDs against morning snapshot to calculate resolution rates.

Usage:
    python run_resolution_agent.py --snapshot afternoon
    python run_resolution_agent.py --snapshot evening
"""

import imaplib
import email
import re
import os
import sys
import argparse
import urllib.request
import openpyxl
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))
GMAIL_USER = "avakash.gupta@wiom.in"
IMAP_SERVER = "imap.gmail.com"
IMAP_PORT = 993
SENDER_FILTER = "kapturecrm.com"
REQUIRED_SUBJECT = "queue wise pending report last 60 days"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def log(msg):
    ts = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S IST")
    print(f"[{ts}] {msg}")


def get_app_password():
    pw = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not pw:
        env_path = os.path.join(SCRIPT_DIR, ".env")
        if os.path.exists(env_path):
            with open(env_path) as f:
                for line in f:
                    if line.startswith("GMAIL_APP_PASSWORD="):
                        pw = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not pw:
        log("ERROR: GMAIL_APP_PASSWORD not set")
        sys.exit(1)
    return pw


def search_emails_for_today():
    """Find all emails for today with the required subject. Returns list of (datetime, download_url) sorted by time."""
    password = get_app_password()

    log("Connecting to Gmail via IMAP...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail.login(GMAIL_USER, password)
    mail.select("inbox")

    today = datetime.now(IST)
    date_str = today.strftime("%d-%b-%Y")

    search_criteria = f'(FROM "{SENDER_FILTER}" SINCE "{date_str}" SUBJECT "pending")'
    log(f"IMAP search: {search_criteria}")

    status, message_ids = mail.search(None, search_criteria)
    if status != "OK" or not message_ids[0]:
        log("No matching emails found for today.")
        mail.logout()
        return []

    ids = message_ids[0].split()
    log(f"Found {len(ids)} candidate email(s)")

    matched = []
    for mid in ids:
        _, msg_data = mail.fetch(mid, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        subject = str(msg.get("Subject", "")).strip()

        if REQUIRED_SUBJECT not in subject.lower():
            continue

        # Parse date
        msg_date = email.utils.parsedate_to_datetime(msg.get("Date", ""))
        msg_date_ist = msg_date.astimezone(IST)

        # Only today's emails
        if msg_date_ist.date() != today.date():
            continue

        # Extract download link
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/html":
                    body = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                    break
        else:
            body = msg.get_payload(decode=True).decode("utf-8", errors="ignore")

        link_match = re.search(r'https://[^\s"<>]+\.xlsx[^\s"<>]*', body)
        if link_match:
            matched.append((msg_date_ist, link_match.group(0), subject))
            log(f"  MATCH: '{subject}' at {msg_date_ist.strftime('%I:%M %p IST')}")

    mail.logout()
    matched.sort(key=lambda x: x[0])  # Sort by time ascending
    return matched


def download_report(url, filename):
    """Download the Excel report."""
    filepath = os.path.join(SCRIPT_DIR, filename)
    log(f"Downloading report to: {filepath}")
    req = urllib.request.Request(url)
    req.add_header("User-Agent", "Mozilla/5.0")
    resp = urllib.request.urlopen(req, timeout=120)
    with open(filepath, "wb") as f:
        f.write(resp.read())
    size_mb = os.path.getsize(filepath) / (1024 * 1024)
    log(f"Downloaded successfully ({size_mb:.1f} MB)")
    return filepath


def extract_ticket_ids_and_l3(xlsx_path):
    """Extract all ticket IDs and their L3 categories from the Excel report."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    ticket_col = headers.index("Ticket No")
    l3_col = headers.index("Disposition Folder Level 3") if "Disposition Folder Level 3" in headers else None

    ticket_ids = []
    l3_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = str(row[ticket_col]).strip() if row[ticket_col] else None
        if tid:
            l3_val = str(row[l3_col]).strip() if l3_col is not None and row[l3_col] else None
            # Exclude Router Pickup tickets from resolution tracking
            if l3_val and l3_val.lower() == "router pickup":
                continue
            ticket_ids.append(tid)
            if l3_val:
                l3_map[tid] = l3_val
    wb.close()
    return ticket_ids, l3_map


def main():
    parser = argparse.ArgumentParser(description="Resolution Agent")
    parser.add_argument("--snapshot", required=True, choices=["morning", "afternoon", "evening"],
                        help="Which snapshot to capture")
    args = parser.parse_args()

    snapshot_type = args.snapshot
    log(f"=== Resolution Agent: {snapshot_type} snapshot ===")

    # Determine which email to pick
    # morning = 1st email, afternoon = 2nd email, evening = 3rd (or latest)
    email_index = {"morning": 0, "afternoon": 1, "evening": 2}[snapshot_type]

    matched = search_emails_for_today()
    if len(matched) <= email_index:
        log(f"Only {len(matched)} email(s) found, need at least {email_index + 1} for {snapshot_type}")
        sys.exit(1)

    _, url, subject = matched[email_index]
    log(f"Using email #{email_index + 1}: '{subject}'")

    today = datetime.now(IST)
    report_date = today.strftime("%Y-%m-%d")
    filename = f"pending_report_{today.strftime('%Y%m%d')}_{snapshot_type}.xlsx"

    # Download
    filepath = download_report(url, filename)

    # Extract ticket IDs
    log("Extracting ticket IDs...")
    ticket_ids, l3_map = extract_ticket_ids_and_l3(filepath)
    log(f"Found {len(ticket_ids)} ticket IDs")

    # Save to DB
    from history_db import save_resolution_snapshot
    save_resolution_snapshot(report_date, snapshot_type, ticket_ids, l3_map if l3_map else None)

    # Clean up the downloaded file (not needed after extracting IDs)
    try:
        os.remove(filepath)
        log(f"Cleaned up: {filename}")
    except Exception:
        pass

    log(f"=== Resolution Agent: {snapshot_type} snapshot complete ===")


if __name__ == "__main__":
    main()
