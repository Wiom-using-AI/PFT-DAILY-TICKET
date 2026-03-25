"""
Backfill Historical Data - March 1 to March 17, 2026
=====================================================
Connects to Gmail via IMAP, finds all daily pending report emails
from March 1-17, downloads each report, and processes through the
full pipeline (save_daily_snapshot, save_full_report, save_category_breakdown).
"""

import imaplib
import email
import re
import os
import sys
import time
from datetime import datetime, timezone, timedelta

# Add parent dir to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from pft_internet_ticket_agent import download_report, filter_internet_tickets, log, IST, DOWNLOAD_DIR
from history_db import save_daily_snapshot, save_full_report, save_category_breakdown

import openpyxl

# --- Configuration ---
GMAIL_USER = "avakash.gupta@wiom.in"
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "hglg cfui wilb rqyt")
IMAP_SERVER = "imap.gmail.com"
IMAP_PORT = 993
SENDER_FILTER = "no-reply-report@kapturecrm.com"
REQUIRED_SUBJECT = "queue wise pending report last 60 days"

# Date range to backfill
START_DATE = datetime(2026, 3, 1, tzinfo=IST)
END_DATE = datetime(2026, 3, 18, tzinfo=IST)  # exclusive — so we get Mar 1-17


def connect_gmail():
    """Connect to Gmail via IMAP and return the mail object."""
    log("Connecting to Gmail via IMAP...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")
    return mail


def find_all_report_emails(mail):
    """
    Search for all pending report emails between START_DATE and END_DATE.
    Returns a list of (email_date_str, email_datetime_ist, download_link) sorted by date.
    """
    since_str = START_DATE.strftime("%d-%b-%Y")  # "01-Mar-2026"
    before_str = END_DATE.strftime("%d-%b-%Y")    # "18-Mar-2026"

    search_criteria = f'(FROM "{SENDER_FILTER}" SINCE "{since_str}" BEFORE "{before_str}" SUBJECT "pending")'
    log(f"IMAP search: {search_criteria}")

    status, message_ids = mail.search(None, search_criteria)
    if status != "OK" or not message_ids[0]:
        log("No matching emails found in date range.")
        return []

    ids = message_ids[0].split()
    log(f"Found {len(ids)} candidate email(s) in date range")

    # Group emails by date, pick earliest per day
    emails_by_date = {}  # date_str -> [(datetime_ist, link, subject)]

    for i, msg_id in enumerate(ids):
        status, msg_data = mail.fetch(msg_id, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(msg_data[0][1])

        # Check subject
        subject = str(msg.get("Subject", "")).strip()
        if REQUIRED_SUBJECT not in subject.lower():
            log(f"  SKIP #{i+1}: wrong subject '{subject}'")
            continue

        # Parse date
        date_str_raw = msg.get("Date", "")
        try:
            msg_date = email.utils.parsedate_to_datetime(date_str_raw)
            if msg_date.tzinfo is None:
                msg_date = msg_date.replace(tzinfo=timezone.utc)
            msg_date_ist = msg_date.astimezone(IST)
        except Exception:
            log(f"  SKIP #{i+1}: could not parse date")
            continue

        # Extract download link from body
        body_text = ""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                if content_type in ("text/plain", "text/html"):
                    payload = part.get_payload(decode=True)
                    if payload:
                        body_text += payload.decode("utf-8", errors="ignore")
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body_text = payload.decode("utf-8", errors="ignore")

        pattern = r'https://storage\.googleapis\.com/kapture_report/EXCEL_Report/[^">\s]+'
        link_match = re.search(pattern, body_text)
        if not link_match:
            log(f"  SKIP #{i+1}: no download link found")
            continue

        date_key = msg_date_ist.strftime("%Y-%m-%d")
        if date_key not in emails_by_date:
            emails_by_date[date_key] = []
        emails_by_date[date_key].append((msg_date_ist, link_match.group(0), subject))
        log(f"  FOUND #{i+1}: {date_key} at {msg_date_ist.strftime('%I:%M %p IST')} — '{subject}'")

    # Pick earliest email per day
    result = []
    for date_key in sorted(emails_by_date.keys()):
        entries = sorted(emails_by_date[date_key], key=lambda x: x[0])
        first_time, first_link, first_subject = entries[0]
        result.append((date_key, first_time, first_link))
        if len(entries) > 1:
            log(f"  {date_key}: Using earliest of {len(entries)} emails ({first_time.strftime('%I:%M %p')})")

    return result


def process_report(report_date_str, report_time_ist, download_link):
    """
    Download a report and process it through the full pipeline:
    1. Download full report xlsx
    2. Filter internet issues tickets
    3. Save daily snapshot (internet issues)
    4. Save full report (all categories with L4)
    5. Save category breakdown
    """
    log(f"\n{'='*60}")
    log(f"PROCESSING: {report_date_str}")
    log(f"{'='*60}")

    # Step 1: Download
    filename = f"pending_report_{report_date_str.replace('-', '')}_morning.xlsx"
    try:
        report_path = download_report(download_link, filename)
    except Exception as e:
        log(f"ERROR downloading {report_date_str}: {e}")
        return False

    # Step 2: Filter internet issues
    try:
        output_path = filter_internet_tickets(report_path)
    except Exception as e:
        log(f"ERROR filtering {report_date_str}: {e}")
        return False

    if not output_path:
        log(f"WARNING: No Internet Issues tickets for {report_date_str}")
        output_path = None

    # Step 3: Count total tickets for total_pending
    try:
        wb_count = openpyxl.load_workbook(report_path, read_only=True)
        ws_count = wb_count.active
        total_pending = sum(1 for _ in ws_count.iter_rows(min_row=2))
        wb_count.close()
    except Exception as e:
        log(f"WARNING: Could not count total: {e}")
        total_pending = None

    # Step 4: Save daily snapshot (internet issues tickets)
    if output_path:
        try:
            snapshot_time = report_time_ist.replace(hour=10, minute=10, second=0, microsecond=0)
            save_daily_snapshot(output_path, report_date_str, snapshot_time, total_pending=total_pending)
            log(f"Daily snapshot saved for {report_date_str}")
        except Exception as e:
            log(f"ERROR saving daily snapshot {report_date_str}: {e}")

    # Step 5: Save full report (all categories with L4)
    try:
        snapshot_time = report_time_ist.replace(hour=10, minute=10, second=0, microsecond=0)
        full_count = save_full_report(report_path, report_date_str, snapshot_time)
        log(f"Full report: {full_count} tickets saved for {report_date_str}")

        # Update total_pending
        import sqlite3
        db_path = os.path.join(SCRIPT_DIR, "ticket_history.db")
        conn = sqlite3.connect(db_path)
        conn.execute(
            "UPDATE daily_summary SET total_pending = ? WHERE report_date = ? AND (total_pending IS NULL OR total_pending = 0)",
            (full_count, report_date_str),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        log(f"ERROR saving full report {report_date_str}: {e}")

    # Step 6: Save category breakdown
    try:
        wb = openpyxl.load_workbook(report_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        l3_idx = headers.index("Disposition Folder Level 3")
        cats = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = str(row[l3_idx] or "Unknown").strip()
            cats[val] = cats.get(val, 0) + 1
        wb.close()
        save_category_breakdown(report_date_str, cats)
        log(f"Category breakdown saved: {len(cats)} categories")
    except Exception as e:
        log(f"ERROR saving categories {report_date_str}: {e}")

    log(f"DONE: {report_date_str}")
    return True


def main():
    log("=" * 60)
    log("BACKFILL HISTORICAL DATA: March 1-17, 2026")
    log("=" * 60)

    # Connect to Gmail
    mail = connect_gmail()

    # Find all report emails
    report_emails = find_all_report_emails(mail)
    mail.logout()

    if not report_emails:
        log("No report emails found in date range. Exiting.")
        return

    log(f"\nFound reports for {len(report_emails)} days:")
    for date_str, time_ist, link in report_emails:
        log(f"  {date_str} ({time_ist.strftime('%I:%M %p IST')})")

    # Process each day
    success = 0
    failed = 0
    for date_str, time_ist, link in report_emails:
        try:
            if process_report(date_str, time_ist, link):
                success += 1
            else:
                failed += 1
        except Exception as e:
            log(f"UNEXPECTED ERROR for {date_str}: {e}")
            failed += 1
        # Small delay between downloads to be nice to the server
        time.sleep(2)

    log(f"\n{'='*60}")
    log(f"BACKFILL COMPLETE")
    log(f"  Success: {success}")
    log(f"  Failed:  {failed}")
    log(f"  Total:   {len(report_emails)} days")
    log(f"{'='*60}")


if __name__ == "__main__":
    main()
