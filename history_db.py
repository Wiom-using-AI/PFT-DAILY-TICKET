"""
History Database - Stores daily ticket snapshots in SQLite
===========================================================
Each day's run saves all Internet Issues tickets into the database.
This enables looking back at any past date (d-1, d-2, d-n).
"""

import sqlite3
import os
from datetime import datetime, timezone, timedelta

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "ticket_history.db")
IST = timezone(timedelta(hours=5, minutes=30))


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Create the database tables if they don't exist."""
    conn = get_connection()
    c = conn.cursor()

    # Daily summary table - one row per report date
    c.execute("""
        CREATE TABLE IF NOT EXISTS daily_summary (
            report_date TEXT PRIMARY KEY,
            report_time TEXT,
            total_pending INTEGER,
            total_internet INTEGER,
            created_today INTEGER,
            critical_gt48h INTEGER,
            bucket_lt4h INTEGER DEFAULT 0,
            bucket_4_12h INTEGER DEFAULT 0,
            bucket_12_24h INTEGER DEFAULT 0,
            bucket_24_36h INTEGER DEFAULT 0,
            bucket_36_48h INTEGER DEFAULT 0,
            bucket_48_72h INTEGER DEFAULT 0,
            bucket_72_120h INTEGER DEFAULT 0,
            bucket_gt120h INTEGER DEFAULT 0,
            queue_partner INTEGER DEFAULT 0,
            queue_cx_high_pain INTEGER DEFAULT 0,
            queue_px_send_wiom INTEGER DEFAULT 0,
            master_total INTEGER DEFAULT 0,
            master_already INTEGER DEFAULT 0,
            master_new INTEGER DEFAULT 0,
            master_new_ids TEXT DEFAULT '',
            master_snapshot_time TEXT,
            inserted_at TEXT
        )
    """)

    # Add columns to existing tables (migration-safe)
    for col, default in [
        ("master_total", 0), ("master_already", 0), ("master_new", 0),
        ("master_new_ids", "''"), ("master_snapshot_time", "NULL"),
        ("category_breakdown", "''"),
        ("queue_category_breakdown", "''"),
    ]:
        try:
            text_cols = ("master_new_ids", "master_snapshot_time", "category_breakdown", "queue_category_breakdown")
            c.execute(f"ALTER TABLE daily_summary ADD COLUMN {col} INTEGER DEFAULT {default}"
                      if col not in text_cols
                      else f"ALTER TABLE daily_summary ADD COLUMN {col} TEXT DEFAULT {default}")
        except Exception:
            pass  # Column already exists

    # Ticket-level detail table - all tickets per day
    c.execute("""
        CREATE TABLE IF NOT EXISTS ticket_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date TEXT NOT NULL,
            ticket_no TEXT,
            created_date TEXT,
            created_time TEXT,
            pending_hours REAL,
            aging_bucket TEXT,
            pending_days INTEGER,
            current_queue TEXT,
            sub_status TEXT,
            status TEXT,
            zone TEXT,
            mapped_partner TEXT,
            city TEXT,
            customer_name TEXT,
            device_id TEXT,
            channel_partner TEXT,
            disposition_l1 TEXT,
            disposition_l2 TEXT,
            disposition_l3 TEXT,
            UNIQUE(report_date, ticket_no)
        )
    """)

    # Index for fast queries
    c.execute("CREATE INDEX IF NOT EXISTS idx_ticket_date ON ticket_history(report_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ticket_no ON ticket_history(ticket_no)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ticket_bucket ON ticket_history(report_date, aging_bucket)")

    # Full report table - ALL tickets from the pending report (all categories)
    c.execute("""
        CREATE TABLE IF NOT EXISTS full_report_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date TEXT NOT NULL,
            ticket_no TEXT,
            created_date TEXT,
            created_time TEXT,
            pending_hours REAL,
            aging_bucket TEXT,
            pending_days INTEGER,
            current_queue TEXT,
            sub_status TEXT,
            status TEXT,
            zone TEXT,
            mapped_partner TEXT,
            city TEXT,
            customer_name TEXT,
            device_id TEXT,
            channel_partner TEXT,
            disposition_l1 TEXT,
            disposition_l2 TEXT,
            disposition_l3 TEXT,
            disposition_l4 TEXT,
            UNIQUE(report_date, ticket_no)
        )
    """)
    # Add disposition_l4 column if it doesn't exist (migration for existing DBs)
    try:
        c.execute("ALTER TABLE full_report_history ADD COLUMN disposition_l4 TEXT")
    except Exception:
        pass  # Column already exists
    c.execute("CREATE INDEX IF NOT EXISTS idx_full_date ON full_report_history(report_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_full_l3 ON full_report_history(report_date, disposition_l3)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_full_l3_bucket ON full_report_history(report_date, disposition_l3, aging_bucket)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_full_l4 ON full_report_history(report_date, disposition_l3, disposition_l4)")

    # Agent attendance table
    c.execute("""
        CREATE TABLE IF NOT EXISTS agent_attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date TEXT NOT NULL,
            agent_name TEXT NOT NULL,
            is_present INTEGER DEFAULT 1,
            UNIQUE(report_date, agent_name)
        )
    """)

    # Agent ticket assignments table
    c.execute("""
        CREATE TABLE IF NOT EXISTS agent_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date TEXT NOT NULL,
            ticket_no TEXT NOT NULL,
            agent_name TEXT NOT NULL,
            assigned_at TEXT,
            created_date TEXT,
            created_time TEXT,
            pending_hours REAL,
            aging_bucket TEXT,
            customer_name TEXT,
            mapped_partner TEXT,
            current_queue TEXT,
            status TEXT,
            sub_status TEXT,
            disposition_l3 TEXT,
            disposition_l4 TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            city TEXT,
            zone TEXT,
            device_id TEXT,
            channel_partner TEXT,
            reopen_count INTEGER DEFAULT 0,
            ground_team_update TEXT DEFAULT '',
            ping_status TEXT DEFAULT '',
            cx_action TEXT DEFAULT '',
            px_call_status TEXT DEFAULT '',
            update_date TEXT DEFAULT '',
            agent_remark TEXT DEFAULT '',
            partner_concern TEXT DEFAULT '',
            original_agent TEXT DEFAULT '',
            is_temp INTEGER DEFAULT 0,
            work_status TEXT DEFAULT 'pending',
            UNIQUE(report_date, ticket_no)
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_assign_date ON agent_assignments(report_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_assign_agent ON agent_assignments(report_date, agent_name)")

    # Migrations: add columns if missing (safe for existing DBs)
    for col_sql in [
        "ALTER TABLE agent_assignments ADD COLUMN partner_concern TEXT DEFAULT ''",
        "ALTER TABLE agent_assignments ADD COLUMN original_agent TEXT DEFAULT ''",
        "ALTER TABLE agent_assignments ADD COLUMN is_temp INTEGER DEFAULT 0",
        "ALTER TABLE agent_assignments ADD COLUMN work_status TEXT DEFAULT 'pending'",
    ]:
        try:
            c.execute(col_sql)
        except Exception:
            pass

    # New tickets cache - stores the filtered CSV at processing time
    # Available for download until 11:59 PM that day, then auto-deleted
    c.execute("""
        CREATE TABLE IF NOT EXISTS new_tickets_cache (
            report_date TEXT PRIMARY KEY,
            csv_data TEXT,
            ticket_count INTEGER,
            created_at TEXT
        )
    """)

    conn.commit()
    conn.close()


# Default agent list
AGENT_LIST = ["Sabir", "Saddam", "Dhananjay", "Nitin", "Deepak", "Vivek", "Sandeep", "Noor"]


def parse_datetime_ist(date_str, time_str):
    """Parse date + time into IST datetime."""
    if not date_str or not time_str:
        return None
    try:
        dt = datetime.strptime(f"{date_str} {time_str}", "%d/%m/%Y %H:%M:%S")
        return dt.replace(tzinfo=IST)
    except (ValueError, TypeError):
        return None


AGING_BUCKETS = [
    ("< 4h", 0, 4),
    ("4h - 12h", 4, 12),
    ("12h - 24h", 12, 24),
    ("24h - 36h", 24, 36),
    ("36h - 48h", 36, 48),
    ("48h - 72h", 48, 72),
    ("72h - 120h", 72, 120),
    ("> 120h", 120, float("inf")),
]


def get_bucket(hours):
    if hours is None:
        return "Unknown"
    for label, low, high in AGING_BUCKETS:
        if low <= hours < high:
            return label
    return "> 120h"


def save_daily_snapshot(filtered_xlsx_path, report_date_str, report_time_ist, total_pending=None):
    """
    Save all tickets from a filtered Internet Issues xlsx into the database.

    Args:
        filtered_xlsx_path: Path to the internet_issues_tickets_*.xlsx file
        report_date_str: Date string like "2026-03-18"
        report_time_ist: datetime object of when report was generated (10 AM IST)
        total_pending: Total tickets in the original pending report (before filtering)
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Check if already saved for this date
    c.execute("SELECT COUNT(*) FROM daily_summary WHERE report_date = ?", (report_date_str,))
    if c.fetchone()[0] > 0:
        print(f"[History] Data for {report_date_str} already exists. Updating...")
        c.execute("DELETE FROM ticket_history WHERE report_date = ?", (report_date_str,))
        c.execute("DELETE FROM daily_summary WHERE report_date = ?", (report_date_str,))

    # Load the filtered xlsx
    wb = openpyxl.load_workbook(filtered_xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    tickets = []
    bucket_counts = {}
    queue_counts = {"Partner": 0, "CX - High Pain": 0, "PX-Send to Wiom": 0}
    critical_count = 0
    created_today_count = 0
    report_date_ddmm = report_time_ist.strftime("%d/%m/%Y")

    for row in ws.iter_rows(min_row=2, values_only=True):
        created_date = row[col.get("Created Date", 1)]
        created_time = row[col.get("Created Time", 2)]
        created_dt = parse_datetime_ist(str(created_date) if created_date else None,
                                         str(created_time) if created_time else None)

        if created_dt and report_time_ist:
            hours = max(0, (report_time_ist - created_dt).total_seconds() / 3600)
        else:
            hours = None

        bucket = get_bucket(hours)
        queue = str(row[col.get("Current Queue Name", 47)] or "Unknown").strip()

        ticket = (
            report_date_str,
            str(row[col.get("Ticket No", 0)] or ""),
            str(created_date or ""),
            str(created_time or ""),
            round(hours, 1) if hours is not None else None,
            bucket,
            row[col.get("Pending No of Days", 63)],
            queue,
            str(row[col.get("Sub Status", 83)] or "Unknown").strip(),
            str(row[col.get("Status", 82)] or "Unknown").strip(),
            str(row[col.get("Zone", 70)] or "").strip(),
            str(row[col.get("Mapped Partner name", 69)] or "").strip(),
            str(row[col.get("City", 72)] or "").strip(),
            str(row[col.get("Customer Name", 65)] or "").strip(),
            str(row[col.get("Device ID", 68)] or "").strip(),
            str(row[col.get("Channel Partner", 67)] or "").strip(),
            str(row[col.get("Disposition Folder Level 1", 39)] or "").strip(),
            str(row[col.get("Disposition Folder Level 2", 40)] or "").strip(),
            str(row[col.get("Disposition Folder Level 3", 41)] or "").strip(),
        )
        tickets.append(ticket)

        # Aggregations
        bucket_counts[bucket] = bucket_counts.get(bucket, 0) + 1
        if queue in queue_counts:
            queue_counts[queue] += 1
        if hours is not None and hours > 48:
            critical_count += 1
        if str(created_date) == report_date_ddmm:
            created_today_count += 1

    wb.close()

    # Insert tickets
    c.executemany("""
        INSERT OR REPLACE INTO ticket_history
        (report_date, ticket_no, created_date, created_time, pending_hours, aging_bucket,
         pending_days, current_queue, sub_status, status, zone, mapped_partner, city,
         customer_name, device_id, channel_partner, disposition_l1, disposition_l2, disposition_l3)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, tickets)

    # Insert daily summary
    c.execute("""
        INSERT INTO daily_summary
        (report_date, report_time, total_pending, total_internet, created_today,
         critical_gt48h, bucket_lt4h, bucket_4_12h, bucket_12_24h, bucket_24_36h,
         bucket_36_48h, bucket_48_72h, bucket_72_120h, bucket_gt120h,
         queue_partner, queue_cx_high_pain, queue_px_send_wiom, inserted_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        report_date_str,
        report_time_ist.strftime("%Y-%m-%d %H:%M:%S"),
        total_pending,
        len(tickets),
        created_today_count,
        critical_count,
        bucket_counts.get("< 4h", 0),
        bucket_counts.get("4h - 12h", 0),
        bucket_counts.get("12h - 24h", 0),
        bucket_counts.get("24h - 36h", 0),
        bucket_counts.get("36h - 48h", 0),
        bucket_counts.get("48h - 72h", 0),
        bucket_counts.get("72h - 120h", 0),
        bucket_counts.get("> 120h", 0),
        queue_counts.get("Partner", 0),
        queue_counts.get("CX - High Pain", 0),
        queue_counts.get("PX-Send to Wiom", 0),
        datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
    ))

    conn.commit()
    conn.close()

    print(f"[History] Saved {len(tickets)} tickets for {report_date_str}")
    return len(tickets)


def save_master_snapshot(report_date_str, master_ids_set, ticket_ids_list):
    """
    Save master sheet comparison at the time of daily run (fixed snapshot).
    Called once when daily agent runs. Does NOT change later.
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    already = [tid for tid in ticket_ids_list if tid in master_ids_set]
    new = [tid for tid in ticket_ids_list if tid not in master_ids_set]

    c.execute("""
        UPDATE daily_summary SET
            master_total = ?,
            master_already = ?,
            master_new = ?,
            master_new_ids = ?,
            master_snapshot_time = ?
        WHERE report_date = ?
    """, (
        len(master_ids_set),
        len(already),
        len(new),
        ",".join(new),
        datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        report_date_str,
    ))

    conn.commit()
    conn.close()
    print(f"[Master] Snapshot saved: {len(already)} existing, {len(new)} new, {len(master_ids_set)} in master")
    return {"already": len(already), "new": len(new), "master_total": len(master_ids_set), "new_ids": new}


def save_category_breakdown(report_date_str, category_counts):
    """
    Save the full Disposition Folder Level 3 breakdown from the FULL pending report.
    category_counts = {"Internet Issues": 1718, "Router Pickup": 10524, ...}
    """
    import json
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE daily_summary SET category_breakdown = ? WHERE report_date = ?",
              (json.dumps(category_counts), report_date_str))
    conn.commit()
    conn.close()
    print(f"[Categories] Saved {len(category_counts)} categories for {report_date_str}")


def save_queue_category_breakdown(report_date_str, queue_breakdown):
    """
    Save the L3 x Queue breakdown from the full pending report.
    queue_breakdown = {"Internet Issues": {"CX - High Pain": 38, "Partner": 1526, ...}, ...}
    """
    import json
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE daily_summary SET queue_category_breakdown = ? WHERE report_date = ?",
              (json.dumps(queue_breakdown), report_date_str))
    conn.commit()
    conn.close()
    print(f"[QueueBreakdown] Saved {len(queue_breakdown)} L3 categories with queue splits for {report_date_str}")


def get_category_breakdown(report_date_str):
    """Get category breakdown for a date."""
    import json
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT category_breakdown FROM daily_summary WHERE report_date = ?", (report_date_str,))
    row = c.fetchone()
    conn.close()
    if row and row["category_breakdown"]:
        try:
            return json.loads(row["category_breakdown"])
        except (json.JSONDecodeError, TypeError):
            return {}
    return {}


def save_full_report(full_xlsx_path, report_date_str, report_time_ist):
    """
    Save all tickets from the full pending report EXCEPT Router Pickup.
    Router Pickup daily totals are kept in category_breakdown JSON.
    This keeps the DB small while preserving detailed data for all other categories.
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Clear existing data for this date
    c.execute("DELETE FROM full_report_history WHERE report_date = ?", (report_date_str,))

    wb = openpyxl.load_workbook(full_xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    tickets = []
    total_count = 0
    skipped_router = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_count += 1
        # Skip Router Pickup — only keep daily count in category_breakdown
        l3_val = str(row[col.get("Disposition Folder Level 3", 41)] or "").strip()
        if l3_val == "Router Pickup":
            skipped_router += 1
            continue

        created_date = row[col.get("Created Date", 1)]
        created_time = row[col.get("Created Time", 2)]
        created_dt = parse_datetime_ist(
            str(created_date) if created_date else None,
            str(created_time) if created_time else None,
        )

        if created_dt and report_time_ist:
            hours = max(0, (report_time_ist - created_dt).total_seconds() / 3600)
        else:
            hours = None

        bucket = get_bucket(hours)
        ticket = (
            report_date_str,
            str(row[col.get("Ticket No", 0)] or ""),
            str(created_date or ""),
            str(created_time or ""),
            round(hours, 1) if hours is not None else None,
            bucket,
            row[col.get("Pending No of Days", 63)] if col.get("Pending No of Days") else None,
            str(row[col.get("Current Queue Name", 47)] or "Unknown").strip(),
            str(row[col.get("Sub Status", 83)] or "Unknown").strip(),
            str(row[col.get("Status", 82)] or "Unknown").strip(),
            str(row[col.get("Zone", 70)] or "").strip(),
            str(row[col.get("Mapped Partner name", 69)] or "").strip(),
            str(row[col.get("City", 72)] or "").strip(),
            str(row[col.get("Customer Name", 65)] or "").strip(),
            str(row[col.get("Device ID", 68)] or "").strip(),
            str(row[col.get("Channel Partner", 67)] or "").strip(),
            str(row[col.get("Disposition Folder Level 1", 39)] or "").strip(),
            str(row[col.get("Disposition Folder Level 2", 40)] or "").strip(),
            l3_val,
            str(row[col.get("Disposition Folder Level 4", 42)] or "").strip(),
        )
        tickets.append(ticket)

    wb.close()

    c.executemany("""
        INSERT OR REPLACE INTO full_report_history
        (report_date, ticket_no, created_date, created_time, pending_hours,
         aging_bucket, pending_days, current_queue, sub_status, status,
         zone, mapped_partner, city, customer_name, device_id, channel_partner,
         disposition_l1, disposition_l2, disposition_l3, disposition_l4)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, tickets)

    conn.commit()
    conn.close()
    print(f"[FullReport] Saved {len(tickets)} tickets (skipped {skipped_router} Router Pickup, {total_count} total) for {report_date_str}")
    return total_count


def save_new_tickets_cache(report_date_str, csv_data, ticket_count):
    """Save the new tickets CSV data at processing time.
    This cached CSV is available for download until 11:59 PM that day."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    now_ist = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        INSERT OR REPLACE INTO new_tickets_cache (report_date, csv_data, ticket_count, created_at)
        VALUES (?, ?, ?, ?)
    """, (report_date_str, csv_data, ticket_count, now_ist))
    conn.commit()
    conn.close()
    print(f"[Cache] Saved {ticket_count} new tickets CSV for {report_date_str}")


def get_new_tickets_cache(report_date_str):
    """Retrieve cached new tickets CSV for a date. Returns (csv_data, ticket_count) or (None, 0)."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT csv_data, ticket_count FROM new_tickets_cache WHERE report_date = ?",
              (report_date_str,))
    row = c.fetchone()
    conn.close()
    if row:
        return row["csv_data"], row["ticket_count"]
    return None, 0


def cleanup_expired_cache():
    """Delete new_tickets_cache entries from previous days (past 11:59 PM)."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    today = datetime.now(IST).strftime("%Y-%m-%d")
    c.execute("DELETE FROM new_tickets_cache WHERE report_date < ?", (today,))
    deleted = c.rowcount
    conn.commit()
    conn.close()
    if deleted:
        print(f"[Cache Cleanup] Removed {deleted} expired new tickets cache entries")


def cleanup_old_data():
    """
    Remove old data to keep the database manageable. Runs after each daily save.
    - full_report_history: 31 days (dashboard chart needs all filters across full range)
    - ticket_history: 7 days (download CSV only needs recent data)
    - agent_assignments: 7 days (operational)
    - Daily summary numbers (daily_summary): kept forever (infinite)
    """
    conn = get_connection()
    c = conn.cursor()
    cutoff_31 = (datetime.now(IST) - timedelta(days=31)).strftime("%Y-%m-%d")
    cutoff_7 = (datetime.now(IST) - timedelta(days=7)).strftime("%Y-%m-%d")

    # Full report history: keep 31 days for dashboard chart filters (queue, L4, bucket)
    c.execute("DELETE FROM full_report_history WHERE report_date < ?", (cutoff_31,))
    del_full = c.rowcount
    # Ticket-level download data: keep 7 days only
    c.execute("DELETE FROM ticket_history WHERE report_date < ?", (cutoff_7,))
    del_tickets = c.rowcount
    c.execute("DELETE FROM agent_assignments WHERE report_date < ?", (cutoff_7,))
    del_assign = c.rowcount

    # Clean expired new tickets cache (previous days)
    today = datetime.now(IST).strftime("%Y-%m-%d")
    c.execute("DELETE FROM new_tickets_cache WHERE report_date < ?", (today,))
    del_cache = c.rowcount

    total_deleted = del_full + del_tickets + del_assign
    if total_deleted > 0 or del_cache > 0:
        conn.commit()
        conn.execute("VACUUM")
        print(f"[Cleanup] Removed {del_full} full_report rows (>31d), {del_tickets} ticket rows + {del_assign} assignments (>7d), {del_cache} expired cache entries")
    else:
        conn.commit()
        print(f"[Cleanup] No old data to remove (full_report cutoff: {cutoff_31}, ticket cutoff: {cutoff_7})")
    conn.close()


# Pivot bucket mapping: combine < 4h and 4h-12h into 0-12h for the pivot display
PIVOT_BUCKETS = [
    ("0-12 hrs", ["< 4h", "4h - 12h"]),
    ("12-24 hrs", ["12h - 24h"]),
    ("24-36 hrs", ["24h - 36h"]),
    ("36-48 hrs", ["36h - 48h"]),
    ("48-72 hrs", ["48h - 72h"]),
    ("72-120 hrs", ["72h - 120h"]),
    (">120 hrs", ["> 120h"]),
]

PIVOT_BUCKET_LABELS = [b[0] for b in PIVOT_BUCKETS]


def get_category_aging_pivot(report_date_str):
    """
    Get a pivot table: Category × Aging Bucket with ticket counts.
    Returns: {
        "categories": ["Internet Issues", "Others", ...],
        "buckets": ["0-12 hrs", "12-24 hrs", ...],
        "data": {"Internet Issues": {"0-12 hrs": 359, "12-24 hrs": 587, ...}, ...},
        "totals_by_cat": {"Internet Issues": 1718, ...},
        "totals_by_bucket": {"0-12 hrs": 364, ...},
        "grand_total": 1868
    }
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT disposition_l3, aging_bucket, COUNT(*) as cnt
        FROM full_report_history
        WHERE report_date = ?
        GROUP BY disposition_l3, aging_bucket
    """, (report_date_str,))
    rows = c.fetchall()
    conn.close()

    if not rows:
        return {}

    # Build pivot
    data = {}
    for row in rows:
        cat = row["disposition_l3"] or "Unknown"
        db_bucket = row["aging_bucket"] or "Unknown"
        cnt = row["cnt"]
        if cat not in data:
            data[cat] = {}
        # Map DB bucket to pivot bucket
        for pivot_label, db_labels in PIVOT_BUCKETS:
            if db_bucket in db_labels:
                data[cat][pivot_label] = data[cat].get(pivot_label, 0) + cnt
                break

    # Sort categories by total descending
    totals_by_cat = {cat: sum(buckets.values()) for cat, buckets in data.items()}
    sorted_cats = sorted(totals_by_cat.keys(), key=lambda c: totals_by_cat[c], reverse=True)

    # Totals by bucket
    totals_by_bucket = {}
    for b_label in PIVOT_BUCKET_LABELS:
        totals_by_bucket[b_label] = sum(data.get(cat, {}).get(b_label, 0) for cat in sorted_cats)

    grand_total = sum(totals_by_cat.values())

    return {
        "categories": sorted_cats,
        "buckets": PIVOT_BUCKET_LABELS,
        "data": data,
        "totals_by_cat": totals_by_cat,
        "totals_by_bucket": totals_by_bucket,
        "grand_total": grand_total,
    }


def get_pivot_l4_breakdown(report_date_str, l3_category, date_from=None, date_to=None):
    """
    Get L4 sub-breakdown for a specific L3 category in pivot table format.
    Returns: { "l4_categories": [...], "data": {"L4Cat": {"0-12 hrs": 5, ...}}, "totals": {...} }
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    if date_from and date_to:
        c.execute("""
            SELECT disposition_l4, aging_bucket, COUNT(*) as cnt
            FROM full_report_history
            WHERE report_date >= ? AND report_date <= ? AND disposition_l3 = ?
            GROUP BY disposition_l4, aging_bucket
        """, (date_from, date_to, l3_category))
    else:
        c.execute("""
            SELECT disposition_l4, aging_bucket, COUNT(*) as cnt
            FROM full_report_history
            WHERE report_date = ? AND disposition_l3 = ?
            GROUP BY disposition_l4, aging_bucket
        """, (report_date_str, l3_category))

    rows = c.fetchall()
    conn.close()

    if not rows:
        return {"l4_categories": [], "data": {}, "totals": {}}

    data = {}
    for row in rows:
        l4 = row["disposition_l4"] or "(No L4)"
        db_bucket = row["aging_bucket"] or "Unknown"
        cnt = row["cnt"]
        if l4 not in data:
            data[l4] = {}
        for pivot_label, db_labels in PIVOT_BUCKETS:
            if db_bucket in db_labels:
                data[l4][pivot_label] = data[l4].get(pivot_label, 0) + cnt
                break

    totals = {l4: sum(buckets.values()) for l4, buckets in data.items()}
    sorted_l4 = sorted(totals.keys(), key=lambda x: totals[x], reverse=True)

    return {
        "l4_categories": sorted_l4,
        "data": data,
        "totals": totals,
    }


def get_full_tickets_by_category_bucket(report_date_str, category=None, bucket=None):
    """
    Get raw ticket data from full report filtered by category and/or aging bucket.
    bucket should be a pivot bucket label like '0-12 hrs'.
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    query = "SELECT * FROM full_report_history WHERE report_date = ?"
    params = [report_date_str]

    if category:
        query += " AND disposition_l3 = ?"
        params.append(category)

    if bucket:
        # Map pivot bucket to DB bucket labels
        db_labels = []
        for pivot_label, labels in PIVOT_BUCKETS:
            if pivot_label == bucket:
                db_labels = labels
                break
        if db_labels:
            placeholders = ",".join("?" * len(db_labels))
            query += f" AND aging_bucket IN ({placeholders})"
            params.extend(db_labels)

    query += " ORDER BY pending_hours DESC"
    c.execute(query, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_available_dates():
    """Return all dates that have data, sorted descending (newest first)."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT report_date FROM daily_summary ORDER BY report_date DESC")
    dates = [row["report_date"] for row in c.fetchall()]
    conn.close()
    return dates


def get_summary_range(date_from, date_to):
    """Get aggregated summary across a date range.
    Returns sums + per-day arrays for computing avg/median/min/max on frontend."""
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Get summed values
    c.execute("""
        SELECT
            COUNT(*) as num_days,
            GROUP_CONCAT(report_date) as dates_included,
            SUM(total_pending) as total_pending,
            SUM(total_internet) as total_internet,
            SUM(created_today) as created_today,
            SUM(critical_gt48h) as critical_gt48h,
            SUM(bucket_lt4h) as bucket_lt4h,
            SUM(bucket_4_12h) as bucket_4_12h,
            SUM(bucket_12_24h) as bucket_12_24h,
            SUM(bucket_24_36h) as bucket_24_36h,
            SUM(bucket_36_48h) as bucket_36_48h,
            SUM(bucket_48_72h) as bucket_48_72h,
            SUM(bucket_72_120h) as bucket_72_120h,
            SUM(bucket_gt120h) as bucket_gt120h,
            SUM(queue_partner) as queue_partner,
            SUM(queue_cx_high_pain) as queue_cx_high_pain,
            SUM(queue_px_send_wiom) as queue_px_send_wiom,
            SUM(master_total) as master_total,
            SUM(master_already) as master_already,
            SUM(master_new) as master_new
        FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
    """, (date_from, date_to))
    row = c.fetchone()
    if not row or row["num_days"] == 0:
        conn.close()
        return None
    result = dict(row)

    # Subtract Router Pickup from total_pending (sum level)
    c.execute("""
        SELECT category_breakdown FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date
    """, (date_from, date_to))
    cb_rows = c.fetchall()
    import json as _json
    total_router = 0
    router_per_day = []
    for cb_row in cb_rows:
        try:
            cb = _json.loads(cb_row["category_breakdown"] or "{}")
            rp = cb.get("Router Pickup", 0)
        except Exception:
            rp = 0
        total_router += rp
        router_per_day.append(rp)
    if result.get("total_pending"):
        result["total_pending"] = (result["total_pending"] or 0) - total_router

    # Get per-day values for frontend aggregation (avg/median/min/max)
    KEYS = ["total_pending", "total_internet", "created_today", "critical_gt48h",
            "queue_partner", "queue_cx_high_pain", "queue_px_send_wiom",
            "bucket_lt4h", "bucket_4_12h", "bucket_12_24h", "bucket_24_36h",
            "bucket_36_48h", "bucket_48_72h", "bucket_72_120h", "bucket_gt120h"]
    c.execute(f"""
        SELECT {', '.join(KEYS)}
        FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date
    """, (date_from, date_to))
    rows = c.fetchall()
    conn.close()

    daily_values = {}
    for k in KEYS:
        daily_values[k] = [dict(r)[k] or 0 for r in rows]
    # Subtract Router Pickup from per-day total_pending values
    for i in range(len(daily_values["total_pending"])):
        if i < len(router_per_day):
            daily_values["total_pending"][i] -= router_per_day[i]
    result["daily_values"] = daily_values

    return result


def get_unique_ticket_counts(date_from, date_to):
    """Get count of unique tickets across a date range.
    Uses the latest appearance of each ticket for queue/status classification."""
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Unique Internet Issues tickets (from ticket_history which only has Internet Issues)
    c.execute("""
        SELECT COUNT(DISTINCT ticket_no) FROM ticket_history
        WHERE report_date >= ? AND report_date <= ?
    """, (date_from, date_to))
    unique_internet = c.fetchone()[0] or 0

    # Unique total tickets (from full_report_history — all categories except Router Pickup)
    c.execute("""
        SELECT COUNT(DISTINCT ticket_no) FROM full_report_history
        WHERE report_date >= ? AND report_date <= ?
    """, (date_from, date_to))
    unique_total = c.fetchone()[0] or 0

    # Critical > 48h unique (tickets that appeared with >48h aging on any day)
    c.execute("""
        SELECT COUNT(DISTINCT ticket_no) FROM ticket_history
        WHERE report_date >= ? AND report_date <= ? AND pending_hours > 48
    """, (date_from, date_to))
    unique_critical = c.fetchone()[0] or 0

    # Queue counts — use latest appearance of each ticket
    c.execute("""
        SELECT current_queue, COUNT(*) FROM (
            SELECT ticket_no, current_queue,
                   ROW_NUMBER() OVER (PARTITION BY ticket_no ORDER BY report_date DESC) as rn
            FROM ticket_history
            WHERE report_date >= ? AND report_date <= ?
        ) WHERE rn = 1
        GROUP BY current_queue
    """, (date_from, date_to))
    queues = {row[0]: row[1] for row in c.fetchall()}

    conn.close()

    return {
        "unique_total": unique_total,
        "unique_internet": unique_internet,
        "unique_critical": unique_critical,
        "unique_partner": queues.get("Partner", 0),
        "unique_cx_high_pain": queues.get("CX - High Pain", 0),
        "unique_px_send_wiom": queues.get("PX-Send to Wiom", 0),
    }


def get_category_aging_pivot_range(date_from, date_to):
    """Get aggregated pivot table across a date range.
    Same structure as get_category_aging_pivot but summed across dates."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT disposition_l3, aging_bucket, COUNT(*) as cnt
        FROM full_report_history
        WHERE report_date >= ? AND report_date <= ?
        GROUP BY disposition_l3, aging_bucket
    """, (date_from, date_to))
    rows = c.fetchall()
    conn.close()

    if not rows:
        return {}

    data = {}
    for row in rows:
        cat = row["disposition_l3"] or "Unknown"
        db_bucket = row["aging_bucket"] or "Unknown"
        cnt = row["cnt"]
        if cat not in data:
            data[cat] = {}
        for pivot_label, db_labels in PIVOT_BUCKETS:
            if db_bucket in db_labels:
                data[cat][pivot_label] = data[cat].get(pivot_label, 0) + cnt
                break

    totals_by_cat = {cat: sum(buckets.values()) for cat, buckets in data.items()}
    sorted_cats = sorted(totals_by_cat.keys(), key=lambda c: totals_by_cat[c], reverse=True)

    totals_by_bucket = {}
    for b_label in PIVOT_BUCKET_LABELS:
        totals_by_bucket[b_label] = sum(data.get(cat, {}).get(b_label, 0) for cat in sorted_cats)

    grand_total = sum(totals_by_cat.values())

    return {
        "categories": sorted_cats,
        "buckets": PIVOT_BUCKET_LABELS,
        "data": data,
        "totals_by_cat": totals_by_cat,
        "totals_by_bucket": totals_by_bucket,
        "grand_total": grand_total,
    }


def get_category_breakdown_range(date_from, date_to):
    """Get aggregated category breakdown across a date range."""
    import json
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT report_date, category_breakdown FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
    """, (date_from, date_to))
    rows = c.fetchall()
    conn.close()

    aggregated = {}
    for row in rows:
        if row["category_breakdown"]:
            try:
                cats = json.loads(row["category_breakdown"])
                for cat, count in cats.items():
                    aggregated[cat] = aggregated.get(cat, 0) + count
            except (json.JSONDecodeError, TypeError):
                pass
    return aggregated


def get_category_daily_trend(date_from, date_to):
    """Return category breakdown for each date in the range.
    Returns: {
        'dates': ['2026-03-18', '2026-03-19', ...],
        'categories': {
            'Router Pickup': {'2026-03-18': 10290, '2026-03-19': 10468, ...},
            'Internet Issues': {'2026-03-18': 1718, '2026-03-19': 1892, ...},
            ...
        }
    }
    """
    import json
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT report_date, category_breakdown FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date ASC
    """, (date_from, date_to))
    rows = c.fetchall()
    conn.close()

    dates = []
    categories = {}
    for row in rows:
        rd = row["report_date"]
        dates.append(rd)
        if row["category_breakdown"]:
            try:
                cats = json.loads(row["category_breakdown"])
                for cat, count in cats.items():
                    if cat not in categories:
                        categories[cat] = {}
                    categories[cat][rd] = count
            except (json.JSONDecodeError, TypeError):
                pass
    return {
        "dates": dates,
        "categories": categories,
    }


def get_aging_daily_trend(date_from, date_to, l3_list=None, l4_list=None):
    """Return aging bucket counts for each date in the range.
    l3_list/l4_list can be comma-separated strings or lists.
    If filters are provided, queries full_report_history.
    """
    conn = get_connection()
    c = conn.cursor()

    bucket_labels = ['< 4h', '4h - 12h', '12h - 24h', '24h - 36h',
                     '36h - 48h', '48h - 72h', '72h - 120h', '> 120h']

    # Parse comma-separated filter values
    l3_vals = []
    l4_vals = []
    if l3_list:
        l3_vals = [x.strip() for x in l3_list.split(",") if x.strip()]
    if l4_list:
        l4_vals = [x.strip() for x in l4_list.split(",") if x.strip()]

    # Get available L3 values
    c.execute("""
        SELECT DISTINCT disposition_l3 FROM full_report_history
        WHERE report_date >= ? AND report_date <= ? AND disposition_l3 != ''
        ORDER BY disposition_l3
    """, (date_from, date_to))
    available_l3 = [r["disposition_l3"] for r in c.fetchall()]

    # Get available L4 values (filtered by selected L3s if any)
    l4_query = """
        SELECT DISTINCT disposition_l4 FROM full_report_history
        WHERE report_date >= ? AND report_date <= ? AND disposition_l4 IS NOT NULL AND disposition_l4 != ''
    """
    l4_params = [date_from, date_to]
    if l3_vals:
        placeholders = ",".join("?" * len(l3_vals))
        l4_query += f" AND disposition_l3 IN ({placeholders})"
        l4_params.extend(l3_vals)
    l4_query += " ORDER BY disposition_l4"
    c.execute(l4_query, l4_params)
    available_l4 = [r["disposition_l4"] for r in c.fetchall()]

    if l3_vals or l4_vals:
        # Query from full_report_history with filters
        query = """
            SELECT report_date, aging_bucket, COUNT(*) as cnt
            FROM full_report_history
            WHERE report_date >= ? AND report_date <= ?
        """
        params = [date_from, date_to]
        if l3_vals:
            placeholders = ",".join("?" * len(l3_vals))
            query += f" AND disposition_l3 IN ({placeholders})"
            params.extend(l3_vals)
        if l4_vals:
            placeholders = ",".join("?" * len(l4_vals))
            query += f" AND disposition_l4 IN ({placeholders})"
            params.extend(l4_vals)
        query += " GROUP BY report_date, aging_bucket ORDER BY report_date ASC"
        c.execute(query, params)
        raw = c.fetchall()

        # Get distinct dates
        c.execute("""
            SELECT DISTINCT report_date FROM full_report_history
            WHERE report_date >= ? AND report_date <= ?
            ORDER BY report_date ASC
        """, (date_from, date_to))
        dates = [r["report_date"] for r in c.fetchall()]

        buckets = {bl: {} for bl in bucket_labels}
        for r in raw:
            ab = r["aging_bucket"]
            if ab in buckets:
                buckets[ab][r["report_date"]] = r["cnt"]
    else:
        # Use pre-aggregated daily_summary (Internet Issues only)
        bucket_keys = [
            ('bucket_lt4h', '< 4h'),
            ('bucket_4_12h', '4h - 12h'),
            ('bucket_12_24h', '12h - 24h'),
            ('bucket_24_36h', '24h - 36h'),
            ('bucket_36_48h', '36h - 48h'),
            ('bucket_48_72h', '48h - 72h'),
            ('bucket_72_120h', '72h - 120h'),
            ('bucket_gt120h', '> 120h'),
        ]

        c.execute("""
            SELECT report_date, bucket_lt4h, bucket_4_12h, bucket_12_24h,
                   bucket_24_36h, bucket_36_48h, bucket_48_72h,
                   bucket_72_120h, bucket_gt120h
            FROM daily_summary
            WHERE report_date >= ? AND report_date <= ?
            ORDER BY report_date ASC
        """, (date_from, date_to))
        rows = c.fetchall()

        dates = []
        buckets = {bl: {} for bl in bucket_labels}
        for row in rows:
            rd = row["report_date"]
            dates.append(rd)
            for bk, bl in bucket_keys:
                buckets[bl][rd] = row[bk] or 0

    conn.close()
    return {
        "dates": dates,
        "buckets": buckets,
        "available_l3": available_l3,
        "available_l4": available_l4,
    }


def get_category_l4_daily_trend(date_from, date_to, l3_category):
    """Return L4 breakdown for a specific L3 category across a date range.
    Returns: {
        'dates': ['2026-03-18', '2026-03-19', ...],
        'l4_categories': {
            'L4 Name': {'2026-03-18': 50, '2026-03-19': 60, ...},
            ...
        },
        'l3_totals': {'2026-03-18': 1244, ...}
    }
    """
    conn = get_connection()
    c = conn.cursor()

    # Get dates in range
    c.execute("""
        SELECT DISTINCT report_date FROM full_report_history
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date ASC
    """, (date_from, date_to))
    dates = [r["report_date"] for r in c.fetchall()]

    # Get L3 totals per date
    c.execute("""
        SELECT report_date, COUNT(*) as cnt
        FROM full_report_history
        WHERE report_date >= ? AND report_date <= ? AND disposition_l3 = ?
        GROUP BY report_date
    """, (date_from, date_to, l3_category))
    l3_totals = {r["report_date"]: r["cnt"] for r in c.fetchall()}

    # Get L4 breakdown per date
    c.execute("""
        SELECT report_date, COALESCE(disposition_l4, '') as l4, COUNT(*) as cnt
        FROM full_report_history
        WHERE report_date >= ? AND report_date <= ? AND disposition_l3 = ?
        GROUP BY report_date, disposition_l4
        ORDER BY cnt DESC
    """, (date_from, date_to, l3_category))

    l4_categories = {}
    for r in c.fetchall():
        l4_name = r["l4"] if r["l4"] else "(No L4)"
        if l4_name not in l4_categories:
            l4_categories[l4_name] = {}
        l4_categories[l4_name][r["report_date"]] = r["cnt"]

    conn.close()
    return {
        "dates": dates,
        "l4_categories": l4_categories,
        "l3_totals": l3_totals,
    }


def get_category_trend_chart(date_from, date_to, bucket_filter=None, l3_filter=None, l4_filter=None, expand_l4=False, queue_filter=None):
    """Return ticket counts grouped by L3 (or L4 if expand_l4=True and single L3) per date.
    Uses category_breakdown from daily_summary for dates without raw data (older than 7 days),
    and full_report_history for dates with raw data (supports bucket/L4/queue filters).
    """
    import json
    conn = get_connection()
    c = conn.cursor()

    # Parse filters
    bucket_vals = [x.strip() for x in bucket_filter.split(",") if x.strip()] if bucket_filter else []
    l3_vals = [x.strip() for x in l3_filter.split(",") if x.strip()] if l3_filter else []
    l4_vals = [x.strip() for x in l4_filter.split(",") if x.strip()] if l4_filter else []
    queue_vals = [x.strip() for x in queue_filter.split(",") if x.strip()] if queue_filter else []

    # Determine grouping: only expand to L4 if explicitly requested
    group_by_l4 = (expand_l4 and len(l3_vals) == 1) or (len(l4_vals) > 0)
    has_advanced_filters = len(bucket_vals) > 0 or group_by_l4 or len(queue_vals) > 0

    # Find which dates have raw data
    c.execute("""
        SELECT DISTINCT report_date FROM full_report_history
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date ASC
    """, (date_from, date_to))
    raw_dates = set(r["report_date"] for r in c.fetchall())

    # Get all dates from daily_summary (infinite retention)
    c.execute("""
        SELECT report_date, category_breakdown, queue_category_breakdown FROM daily_summary
        WHERE report_date >= ? AND report_date <= ?
        ORDER BY report_date ASC
    """, (date_from, date_to))
    summary_rows = c.fetchall()

    categories = {}
    dates = []

    for row in summary_rows:
        rd = row["report_date"]
        dates.append(rd)

        if rd in raw_dates and (has_advanced_filters or group_by_l4):
            # Use raw data for this date (supports bucket/L4/queue filters)
            continue  # Will be filled below
        elif len(queue_vals) > 0 and rd not in raw_dates:
            # Queue filter active but no raw data — use queue_category_breakdown from summary
            qbd = json.loads(row["queue_category_breakdown"]) if row["queue_category_breakdown"] else {}
            for cat, queue_counts in qbd.items():
                if l3_vals and cat not in l3_vals:
                    continue
                # Sum counts for selected queues only
                total = sum(queue_counts.get(q, 0) for q in queue_vals)
                if total > 0:
                    if cat not in categories:
                        categories[cat] = {}
                    categories[cat][rd] = total
        else:
            # Use category_breakdown from daily_summary (L3 level only)
            breakdown = json.loads(row["category_breakdown"]) if row["category_breakdown"] else {}
            for cat, cnt in breakdown.items():
                if l3_vals and cat not in l3_vals:
                    continue
                if cat not in categories:
                    categories[cat] = {}
                categories[cat][rd] = cnt

    # For dates with raw data and advanced filters, query full_report_history
    dates_needing_raw = [d for d in dates if d in raw_dates and (has_advanced_filters or group_by_l4)]
    if dates_needing_raw:
        group_col = "COALESCE(disposition_l4, '(No L4)')" if group_by_l4 else "disposition_l3"
        placeholders_dates = ",".join("?" * len(dates_needing_raw))
        query = f"""
            SELECT report_date, {group_col} as category, COUNT(*) as cnt
            FROM full_report_history
            WHERE report_date IN ({placeholders_dates})
              AND disposition_l3 IS NOT NULL AND disposition_l3 != ''
        """
        params = list(dates_needing_raw)

        if bucket_vals:
            ph = ",".join("?" * len(bucket_vals))
            query += f" AND aging_bucket IN ({ph})"
            params.extend(bucket_vals)
        if l3_vals:
            ph = ",".join("?" * len(l3_vals))
            query += f" AND disposition_l3 IN ({ph})"
            params.extend(l3_vals)
        if l4_vals:
            ph = ",".join("?" * len(l4_vals))
            query += f" AND disposition_l4 IN ({ph})"
            params.extend(l4_vals)
        if queue_vals:
            ph = ",".join("?" * len(queue_vals))
            query += f" AND current_queue IN ({ph})"
            params.extend(queue_vals)

        query += " GROUP BY report_date, category ORDER BY report_date ASC"
        c.execute(query, params)

        for r in c.fetchall():
            cat = r["category"] if r["category"] else "(Unknown)"
            if cat not in categories:
                categories[cat] = {}
            categories[cat][r["report_date"]] = r["cnt"]

    # Also fill raw dates that don't need advanced filters but weren't in summary
    # (edge case: date in raw but not in summary)
    for rd in raw_dates:
        if rd not in dates:
            dates.append(rd)

    dates.sort()

    # Sort categories by total count (descending)
    sorted_cats = sorted(categories.keys(), key=lambda k: sum(categories[k].values()), reverse=True)
    categories_sorted = {k: categories[k] for k in sorted_cats}

    # Get available L3 from both sources
    all_l3 = set()
    for row in summary_rows:
        breakdown = json.loads(row["category_breakdown"]) if row["category_breakdown"] else {}
        all_l3.update(breakdown.keys())
    # Also from raw data
    c.execute("""
        SELECT DISTINCT disposition_l3 FROM full_report_history
        WHERE report_date >= ? AND report_date <= ? AND disposition_l3 IS NOT NULL AND disposition_l3 != ''
    """, (date_from, date_to))
    all_l3.update(r["disposition_l3"] for r in c.fetchall())
    available_l3 = sorted(all_l3)

    # Available L4 (only from raw data since summary doesn't store L4)
    available_l4 = []
    if raw_dates:
        l4_query = """
            SELECT DISTINCT disposition_l4 FROM full_report_history
            WHERE report_date >= ? AND report_date <= ? AND disposition_l4 IS NOT NULL AND disposition_l4 != ''
        """
        l4_params = [date_from, date_to]
        if l3_vals:
            ph = ",".join("?" * len(l3_vals))
            l4_query += f" AND disposition_l3 IN ({ph})"
            l4_params.extend(l3_vals)
        l4_query += " ORDER BY disposition_l4"
        c.execute(l4_query, l4_params)
        available_l4 = [r["disposition_l4"] for r in c.fetchall()]

    # Available queues (from raw data + queue_category_breakdown in summaries)
    all_queues = set()
    if raw_dates:
        q_query = """
            SELECT DISTINCT current_queue FROM full_report_history
            WHERE report_date >= ? AND report_date <= ? AND current_queue IS NOT NULL AND current_queue != ''
        """
        q_params = [date_from, date_to]
        if l3_vals:
            ph = ",".join("?" * len(l3_vals))
            q_query += f" AND disposition_l3 IN ({ph})"
            q_params.extend(l3_vals)
        q_query += " ORDER BY current_queue"
        c.execute(q_query, q_params)
        all_queues.update(r["current_queue"] for r in c.fetchall())
    # Also gather queue names from summary queue_category_breakdown
    for row in summary_rows:
        qbd = json.loads(row["queue_category_breakdown"]) if row["queue_category_breakdown"] else {}
        for cat, queue_counts in qbd.items():
            if l3_vals and cat not in l3_vals:
                continue
            all_queues.update(queue_counts.keys())
    available_queues = sorted(all_queues)

    conn.close()
    return {
        "dates": dates,
        "categories": categories_sorted,
        "group_by": "l4" if group_by_l4 else "l3",
        "available_l3": available_l3,
        "available_l4": available_l4,
        "available_queues": available_queues,
        "raw_data_dates": sorted(raw_dates),
    }


def get_tickets_for_download(report_date_str, l3_category=None, l4_category=None):
    """Get raw tickets for download, optionally filtered by L3 and/or L4 category."""
    init_db()
    conn = get_connection()
    c = conn.cursor()

    query = "SELECT * FROM full_report_history WHERE report_date = ?"
    params = [report_date_str]

    if l3_category:
        query += " AND disposition_l3 = ?"
        params.append(l3_category)

    if l4_category:
        if l4_category == "(No L4)":
            query += " AND (disposition_l4 IS NULL OR disposition_l4 = '')"
        else:
            query += " AND disposition_l4 = ?"
            params.append(l4_category)

    query += " ORDER BY pending_hours DESC"
    c.execute(query, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_daily_summary(report_date):
    """Get summary for a specific date."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM daily_summary WHERE report_date = ?", (report_date,))
    row = c.fetchone()
    conn.close()
    return dict(row) if row else None


def get_all_summaries():
    """Get all daily summaries for trend charts."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM daily_summary ORDER BY report_date ASC")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_tickets_for_date(report_date):
    """Get all ticket details for a specific date."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM ticket_history WHERE report_date = ? ORDER BY pending_hours DESC", (report_date,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_all_tickets_for_date(report_date):
    """Get ALL ticket columns for a specific date (for CSV download)."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM ticket_history WHERE report_date = ? ORDER BY pending_hours DESC", (report_date,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_ticket_trail(ticket_no):
    """Get the history of a specific ticket across all dates it appeared."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT report_date, pending_hours, aging_bucket, current_queue, sub_status, status
        FROM ticket_history
        WHERE ticket_no = ?
        ORDER BY report_date ASC
    """, (ticket_no,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


# ==================== AGENT FUNCTIONS ====================

def get_agent_dates():
    """Return all dates that have agent assignments, sorted descending."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT report_date FROM agent_assignments ORDER BY report_date DESC")
    dates = [row["report_date"] for row in c.fetchall()]
    conn.close()
    return dates


def save_attendance(report_date_str, present_agents):
    """Save which agents are present for a given date."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    # Clear existing attendance for this date
    c.execute("DELETE FROM agent_attendance WHERE report_date = ?", (report_date_str,))
    for agent in AGENT_LIST:
        is_present = 1 if agent in present_agents else 0
        c.execute("INSERT INTO agent_attendance (report_date, agent_name, is_present) VALUES (?,?,?)",
                  (report_date_str, agent, is_present))
    conn.commit()
    conn.close()
    return {"present": present_agents, "total": len(AGENT_LIST)}


def get_attendance(report_date_str):
    """Get attendance for a date. Returns dict of agent_name -> is_present."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT agent_name, is_present FROM agent_attendance WHERE report_date = ?", (report_date_str,))
    rows = c.fetchall()
    conn.close()
    if rows:
        return {row["agent_name"]: bool(row["is_present"]) for row in rows}
    # Default: all present
    return {agent: True for agent in AGENT_LIST}


def assign_tickets_round_robin(report_date_str, present_agents=None):
    """
    Assign tickets for the given date using round-robin:
    1. Today's new tickets → round-robin to present agents
    2. Absent agents' PENDING tickets from previous days → temp-redistribute to present agents
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Check if today's new tickets are already assigned
    c.execute("SELECT COUNT(*) FROM agent_assignments WHERE report_date = ? AND is_temp = 0", (report_date_str,))
    existing = c.fetchone()[0]
    if existing > 0:
        conn.close()
        return {"status": "already_assigned", "count": existing}

    # Get present agents
    if not present_agents:
        attendance = get_attendance(report_date_str)
        present_agents = [a for a, p in attendance.items() if p]
    if not present_agents:
        conn.close()
        return {"status": "error", "message": "No agents marked as present"}

    absent_agents = [a for a in AGENT_LIST if a not in present_agents]

    # ---- Part 1: Assign today's new tickets (only Ticket Pending) ----
    c.execute("SELECT master_new_ids FROM daily_summary WHERE report_date = ?", (report_date_str,))
    row = c.fetchone()
    new_ids = set()
    if row and row["master_new_ids"]:
        new_ids = set(x.strip() for x in row["master_new_ids"].split(",") if x.strip())

    if new_ids:
        placeholders = ",".join("?" * len(new_ids))
        c.execute(f"""
            SELECT * FROM ticket_history
            WHERE report_date = ? AND ticket_no IN ({placeholders})
            ORDER BY pending_hours DESC
        """, [report_date_str] + list(new_ids))
    else:
        c.execute("SELECT * FROM ticket_history WHERE report_date = ? ORDER BY pending_hours DESC",
                  (report_date_str,))
    tickets = [dict(r) for r in c.fetchall()]

    new_assigned = 0
    now_str = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")

    if tickets:
        assignments = []
        for i, ticket in enumerate(tickets):
            agent = present_agents[i % len(present_agents)]
            assignments.append((
                report_date_str, ticket["ticket_no"], agent, now_str,
                ticket.get("created_date", ""), ticket.get("created_time", ""),
                ticket.get("pending_hours"), ticket.get("aging_bucket", ""),
                ticket.get("customer_name", ""), ticket.get("mapped_partner", ""),
                ticket.get("current_queue", ""), ticket.get("status", ""),
                ticket.get("sub_status", ""), ticket.get("disposition_l3", ""),
                "", "",  # disposition_l4, phone
                ticket.get("city", ""), ticket.get("zone", ""),
                ticket.get("device_id", ""), ticket.get("channel_partner", ""),
                agent,  # original_agent = same as agent (it's their own ticket)
                0,      # is_temp = 0
                "pending",  # work_status
            ))

        c.executemany("""
            INSERT OR REPLACE INTO agent_assignments
            (report_date, ticket_no, agent_name, assigned_at,
             created_date, created_time, pending_hours, aging_bucket,
             customer_name, mapped_partner, current_queue, status, sub_status,
             disposition_l3, disposition_l4, phone, city, zone, device_id, channel_partner,
             original_agent, is_temp, work_status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, assignments)
        new_assigned = len(assignments)

    # ---- Part 2: Redistribute absent agents' PENDING tickets ----
    # Only tickets with Kapture status != 'Ticket Closed' and work_status = 'pending'
    redistributed = 0
    if absent_agents:
        abs_placeholders = ",".join("?" * len(absent_agents))
        c.execute(f"""
            SELECT * FROM agent_assignments
            WHERE original_agent IN ({abs_placeholders})
              AND work_status = 'pending'
              AND is_temp = 0
              AND report_date != ?
              AND COALESCE(status, '') != 'Ticket Closed'
            ORDER BY pending_hours DESC
        """, absent_agents + [report_date_str])
        pending_tickets = [dict(r) for r in c.fetchall()]

        if pending_tickets:
            for i, ticket in enumerate(pending_tickets):
                temp_agent = present_agents[i % len(present_agents)]
                c.execute("""
                    UPDATE agent_assignments
                    SET agent_name = ?, is_temp = 1, assigned_at = ?
                    WHERE report_date = ? AND ticket_no = ?
                """, (temp_agent, now_str, ticket["report_date"], ticket["ticket_no"]))
            redistributed = len(pending_tickets)

    conn.commit()
    conn.close()

    # Count per agent across both new + redistributed
    result_counts = {}
    for a in present_agents:
        result_counts[a] = 0
    if tickets:
        for i, _ in enumerate(tickets):
            agent = present_agents[i % len(present_agents)]
            result_counts[agent] = result_counts.get(agent, 0) + 1

    return {
        "status": "assigned",
        "total": new_assigned,
        "redistributed": redistributed,
        "absent_agents": absent_agents,
        "agents": len(present_agents),
        "per_agent": result_counts,
    }


def get_agent_assignments(report_date_str, agent_name=None):
    """Get ticket assignments for a date, optionally filtered by agent."""
    init_db()
    conn = get_connection()
    c = conn.cursor()
    if agent_name:
        c.execute("""SELECT * FROM agent_assignments
                     WHERE report_date = ? AND agent_name = ?
                     ORDER BY pending_hours DESC""",
                  (report_date_str, agent_name))
    else:
        c.execute("""SELECT * FROM agent_assignments
                     WHERE report_date = ? ORDER BY agent_name, pending_hours DESC""",
                  (report_date_str,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def get_agent_active_tickets(report_date_str, agent_name=None):
    """
    Get ALL tickets an agent is currently responsible for:
    - Own tickets from today (report_date, is_temp=0)
    - Temp-redistributed tickets from past dates (is_temp=1, agent_name=current holder)
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()

    if agent_name:
        # Today's own tickets for this agent
        c.execute("""SELECT *, 'own' as ticket_type FROM agent_assignments
                     WHERE report_date = ? AND agent_name = ? AND is_temp = 0
                     ORDER BY pending_hours DESC""",
                  (report_date_str, agent_name))
        own = [dict(r) for r in c.fetchall()]

        # Temp tickets currently held by this agent (from any date)
        c.execute("""SELECT *, 'temp' as ticket_type FROM agent_assignments
                     WHERE agent_name = ? AND is_temp = 1 AND work_status = 'pending'
                     ORDER BY pending_hours DESC""",
                  (agent_name,))
        temp = [dict(r) for r in c.fetchall()]
    else:
        # All today's own tickets
        c.execute("""SELECT *, 'own' as ticket_type FROM agent_assignments
                     WHERE report_date = ? AND is_temp = 0
                     ORDER BY agent_name, pending_hours DESC""",
                  (report_date_str,))
        own = [dict(r) for r in c.fetchall()]

        # All temp tickets currently active
        c.execute("""SELECT *, 'temp' as ticket_type FROM agent_assignments
                     WHERE is_temp = 1 AND work_status = 'pending'
                     ORDER BY agent_name, pending_hours DESC""",
                  ())
        temp = [dict(r) for r in c.fetchall()]

    conn.close()
    return own + temp


def get_agent_summary(report_date_str):
    """Get assignment count per agent for a date, including cross-date temp tickets."""
    init_db()
    conn = get_connection()
    c = conn.cursor()

    # Today's own tickets (assigned on this date, not temp)
    c.execute("""SELECT agent_name, COUNT(*) as count
                 FROM agent_assignments WHERE report_date = ? AND is_temp = 0
                 GROUP BY agent_name ORDER BY agent_name""",
              (report_date_str,))
    own = {row["agent_name"]: row["count"] for row in c.fetchall()}

    # Temp tickets currently held by each agent (from any date)
    c.execute("""SELECT agent_name, COUNT(*) as count
                 FROM agent_assignments WHERE is_temp = 1 AND work_status = 'pending'
                 GROUP BY agent_name""")
    temp = {row["agent_name"]: row["count"] for row in c.fetchall()}

    # Pending count per agent (own, across all dates)
    c.execute("""SELECT original_agent, COUNT(*) as count
                 FROM agent_assignments WHERE work_status = 'pending' AND is_temp = 0
                 GROUP BY original_agent""")
    pending = {row["original_agent"]: row["count"] for row in c.fetchall()}

    # Completed count per agent (own, across all dates)
    c.execute("""SELECT original_agent, COUNT(*) as count
                 FROM agent_assignments WHERE work_status = 'completed' AND is_temp = 0
                 GROUP BY original_agent""")
    completed = {row["original_agent"]: row["count"] for row in c.fetchall()}

    conn.close()

    result = {}
    for agent in AGENT_LIST:
        result[agent] = {
            "own_today": own.get(agent, 0),
            "temp_holding": temp.get(agent, 0),
            "total_today": own.get(agent, 0) + temp.get(agent, 0),
            "pending_all": pending.get(agent, 0),
            "completed_all": completed.get(agent, 0),
        }
    return result


def update_agent_ticket(report_date_str, ticket_no, updates):
    """Update agent work fields for a specific ticket."""
    init_db()
    allowed = {"ground_team_update", "ping_status", "cx_action",
               "px_call_status", "update_date", "agent_remark", "partner_concern",
               "work_status"}
    filtered = {k: v for k, v in updates.items() if k in allowed}
    if not filtered:
        return False
    conn = get_connection()
    c = conn.cursor()
    sets = ", ".join(f"{k} = ?" for k in filtered)
    vals = list(filtered.values()) + [report_date_str, ticket_no]
    c.execute(f"UPDATE agent_assignments SET {sets} WHERE report_date = ? AND ticket_no = ?", vals)
    conn.commit()
    conn.close()
    return True


def reassign_tickets(report_date_str, present_agents):
    """
    Re-assign tickets for a date with updated attendance:
    1. Today's OWN tickets (is_temp=0, report_date=today) → re-distribute among present agents
    2. Absent agents' pending tickets from past dates → temp-redistribute to present
    3. Returning agents → reclaim their original pending tickets that were temp-assigned
    """
    init_db()
    conn = get_connection()
    c = conn.cursor()
    now_str = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")

    absent_agents = [a for a in AGENT_LIST if a not in present_agents]
    reclaimed = 0
    redistributed = 0

    # ---- Step 1: Reclaim tickets for RETURNING agents ----
    # If an agent is now present, their original pending tickets (not closed)
    # that were temp-assigned to someone else go back to them
    for agent in present_agents:
        c.execute("""
            UPDATE agent_assignments
            SET agent_name = original_agent, is_temp = 0, assigned_at = ?
            WHERE original_agent = ? AND is_temp = 1 AND work_status = 'pending'
              AND COALESCE(status, '') != 'Ticket Closed'
        """, (now_str, agent))
        reclaimed += c.rowcount

    # ---- Step 2: Re-distribute today's own tickets among present agents ----
    # Delete today's assignments and re-do round-robin
    c.execute("DELETE FROM agent_assignments WHERE report_date = ? AND is_temp = 0", (report_date_str,))
    conn.commit()

    # Get ticket data again
    c.execute("SELECT master_new_ids FROM daily_summary WHERE report_date = ?", (report_date_str,))
    row = c.fetchone()
    new_ids = set()
    if row and row["master_new_ids"]:
        new_ids = set(x.strip() for x in row["master_new_ids"].split(",") if x.strip())

    if new_ids:
        placeholders = ",".join("?" * len(new_ids))
        c.execute(f"""
            SELECT * FROM ticket_history
            WHERE report_date = ? AND ticket_no IN ({placeholders})
            ORDER BY pending_hours DESC
        """, [report_date_str] + list(new_ids))
    else:
        c.execute("SELECT * FROM ticket_history WHERE report_date = ? ORDER BY pending_hours DESC",
                  (report_date_str,))
    tickets = [dict(r) for r in c.fetchall()]

    new_assigned = 0
    if tickets:
        assignments = []
        for i, ticket in enumerate(tickets):
            agent = present_agents[i % len(present_agents)]
            assignments.append((
                report_date_str, ticket["ticket_no"], agent, now_str,
                ticket.get("created_date", ""), ticket.get("created_time", ""),
                ticket.get("pending_hours"), ticket.get("aging_bucket", ""),
                ticket.get("customer_name", ""), ticket.get("mapped_partner", ""),
                ticket.get("current_queue", ""), ticket.get("status", ""),
                ticket.get("sub_status", ""), ticket.get("disposition_l3", ""),
                "", "",
                ticket.get("city", ""), ticket.get("zone", ""),
                ticket.get("device_id", ""), ticket.get("channel_partner", ""),
                agent, 0, "pending",
            ))

        c.executemany("""
            INSERT OR REPLACE INTO agent_assignments
            (report_date, ticket_no, agent_name, assigned_at,
             created_date, created_time, pending_hours, aging_bucket,
             customer_name, mapped_partner, current_queue, status, sub_status,
             disposition_l3, disposition_l4, phone, city, zone, device_id, channel_partner,
             original_agent, is_temp, work_status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, assignments)
        new_assigned = len(assignments)

    # ---- Step 3: Redistribute absent agents' pending tickets (only non-closed) ----
    if absent_agents:
        abs_placeholders = ",".join("?" * len(absent_agents))
        c.execute(f"""
            SELECT * FROM agent_assignments
            WHERE original_agent IN ({abs_placeholders})
              AND work_status = 'pending'
              AND is_temp = 0
              AND report_date != ?
              AND COALESCE(status, '') != 'Ticket Closed'
            ORDER BY pending_hours DESC
        """, absent_agents + [report_date_str])
        pending_tickets = [dict(r) for r in c.fetchall()]

        if pending_tickets:
            for i, ticket in enumerate(pending_tickets):
                temp_agent = present_agents[i % len(present_agents)]
                c.execute("""
                    UPDATE agent_assignments
                    SET agent_name = ?, is_temp = 1, assigned_at = ?
                    WHERE report_date = ? AND ticket_no = ?
                """, (temp_agent, now_str, ticket["report_date"], ticket["ticket_no"]))
            redistributed = len(pending_tickets)

    conn.commit()
    conn.close()

    return {
        "status": "assigned",
        "total": new_assigned,
        "redistributed": redistributed,
        "reclaimed": reclaimed,
        "absent_agents": absent_agents,
        "agents": len(present_agents),
    }


if __name__ == "__main__":
    init_db()
    print(f"Database initialized at: {DB_PATH}")
    dates = get_available_dates()
    if dates:
        print(f"Available dates: {', '.join(dates)}")
    else:
        print("No data yet. Run the daily agent to start collecting.")
