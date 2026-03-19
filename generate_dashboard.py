"""
PFT Dashboard Generator
========================
Generates an interactive HTML dashboard from the filtered Internet Issues report.
Shows aging buckets (in hours), queue-wise split, zone-wise, status breakdowns, and more.
"""

import os
import sys
import json
import glob
from datetime import datetime, timezone, timedelta
from collections import Counter, defaultdict

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
IST = timezone(timedelta(hours=5, minutes=30))

# Default report reference time — will be overridden by actual email time
REPORT_TIME_IST = None

# Aging buckets in hours
AGING_BUCKETS = [
    ("4h - 12h", 4, 12),
    ("12h - 24h", 12, 24),
    ("24h - 36h", 24, 36),
    ("36h - 48h", 36, 48),
    ("48h - 72h", 48, 72),
    ("72h - 120h", 72, 120),
    ("> 120h", 120, float("inf")),
]

BUCKET_COLORS = [
    "#22c55e",  # green
    "#84cc16",  # lime
    "#eab308",  # yellow
    "#f97316",  # orange
    "#ef4444",  # red
    "#dc2626",  # dark red
    "#991b1b",  # very dark red
]


def parse_datetime(date_str, time_str):
    """Parse Created Date + Created Time into a datetime object."""
    if not date_str or not time_str:
        return None
    try:
        dt_str = f"{date_str} {time_str}"
        # Format: DD/MM/YYYY HH:MM:SS
        dt = datetime.strptime(dt_str, "%d/%m/%Y %H:%M:%S")
        return dt.replace(tzinfo=IST)
    except (ValueError, TypeError):
        return None


def calc_pending_hours(created_dt, reference_time):
    """Calculate hours between created datetime and the report snapshot time."""
    if not created_dt or not reference_time:
        return None
    diff = reference_time - created_dt
    total_hours = diff.total_seconds() / 3600
    return total_hours if total_hours >= 0 else 0


def get_bucket(hours):
    """Assign an aging bucket label based on pending hours."""
    if hours is None or hours < 4:
        return "< 4h"
    for label, low, high in AGING_BUCKETS:
        if low <= hours < high:
            return label
    return "> 120h"


def detect_report_time(filepath):
    """Detect the report snapshot time from the filename or file metadata.

    The 10 AM report filename pattern: pending_report_YYYYMMDD_morning.xlsx
    or pending_report_YYYYMMDD_HHMM.xlsx
    Falls back to file modification time.
    """
    basename = os.path.basename(filepath)

    # Try to extract date from filename like "pending_report_20260318_morning.xlsx"
    import re
    date_match = re.search(r'(\d{8})', basename)
    if date_match:
        date_str = date_match.group(1)
        try:
            report_date = datetime.strptime(date_str, "%Y%m%d")
            # The 10 AM IST report — use 10:10 AM as the reference
            return report_date.replace(hour=10, minute=10, tzinfo=IST)
        except ValueError:
            pass

    # Fallback: use file modification time
    mtime = os.path.getmtime(filepath)
    return datetime.fromtimestamp(mtime, tz=IST)


def load_and_analyze(filepath, report_time=None):
    """Load the filtered Internet Issues xlsx and compute all dashboard metrics."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Column indices
    col = {h: i for i, h in enumerate(headers) if h}

    # Determine the report snapshot time
    if report_time is None:
        report_time = REPORT_TIME_IST or detect_report_time(filepath)

    print(f"Report reference time: {report_time.strftime('%d %b %Y, %I:%M %p IST')}")

    tickets = []
    bucket_counts = Counter()
    queue_counts = Counter()
    sub_status_counts = Counter()
    zone_counts = Counter()
    partner_counts = Counter()
    city_counts = Counter()
    daily_created = Counter()
    bucket_by_queue = defaultdict(Counter)

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticket = {
            "ticket_no": row[col.get("Ticket No", 0)],
            "created_date": row[col.get("Created Date", 1)],
            "created_time": row[col.get("Created Time", 2)],
            "pending_days": row[col.get("Pending No of Days", 63)],
            "queue": str(row[col.get("Current Queue Name", 47)] or "Unknown").strip(),
            "sub_status": str(row[col.get("Sub Status", 83)] or "Unknown").strip(),
            "status": str(row[col.get("Status", 82)] or "Unknown").strip(),
            "zone": str(row[col.get("Zone", 70)] or "Unknown").strip(),
            "partner": str(row[col.get("Mapped Partner name", 69)] or "Unknown").strip(),
            "city": str(row[col.get("City", 72)] or "Unknown").strip(),
            "customer": str(row[col.get("Customer Name", 65)] or "").strip(),
            "device_id": str(row[col.get("Device ID", 68)] or "").strip(),
        }

        # Calculate pending hours relative to report snapshot time (10 AM)
        created_dt = parse_datetime(ticket["created_date"], ticket["created_time"])
        hours = calc_pending_hours(created_dt, report_time)
        ticket["pending_hours"] = round(hours, 1) if hours else None
        bucket = get_bucket(hours)
        ticket["bucket"] = bucket

        tickets.append(ticket)

        # Aggregations
        bucket_counts[bucket] += 1
        queue_counts[ticket["queue"]] += 1
        sub_status_counts[ticket["sub_status"]] += 1
        if ticket["zone"] and ticket["zone"] != "Unknown":
            zone_key = ticket["zone"].split(",")[0].strip()
            zone_counts[zone_key] += 1
        partner_counts[ticket["partner"]] += 1
        if ticket["city"] and ticket["city"].strip():
            city_counts[ticket["city"].strip()] += 1
        if ticket["created_date"]:
            daily_created[ticket["created_date"]] += 1
        bucket_by_queue[ticket["queue"]][bucket] += 1

    wb.close()

    # Sort buckets in order
    all_bucket_labels = ["< 4h"] + [b[0] for b in AGING_BUCKETS]
    ordered_buckets = [(label, bucket_counts.get(label, 0)) for label in all_bucket_labels]

    # Critical tickets (> 48h)
    critical_count = sum(1 for t in tickets if t["pending_hours"] and t["pending_hours"] > 48)
    # Uploaded today (created on the report date)
    report_date_str = report_time.strftime("%d/%m/%Y")
    uploaded_today = sum(1 for t in tickets if t["created_date"] == report_date_str)

    return {
        "total": len(tickets),
        "critical": critical_count,
        "uploaded_today": uploaded_today,
        "buckets": ordered_buckets,
        "queues": queue_counts.most_common(),
        "sub_statuses": sub_status_counts.most_common(),
        "zones": zone_counts.most_common(20),
        "partners": partner_counts.most_common(20),
        "daily_created": sorted(daily_created.items()),
        "bucket_by_queue": {q: {b: bucket_by_queue[q].get(b, 0) for b in all_bucket_labels} for q in queue_counts},
        "tickets": tickets,
        "report_time": report_time.strftime("%d %b %Y, %I:%M %p IST"),
        "all_bucket_labels": all_bucket_labels,
    }


def generate_html(data, master_sheet_url, output_path):
    """Generate the full HTML dashboard."""
    bucket_labels = json.dumps(data["all_bucket_labels"])
    bucket_values = json.dumps([dict(data["buckets"]).get(b, 0) for b in data["all_bucket_labels"]])
    bucket_colors_json = json.dumps(["#3b82f6"] + BUCKET_COLORS)

    queue_labels = json.dumps([q[0] for q in data["queues"]])
    queue_values = json.dumps([q[1] for q in data["queues"]])

    sub_status_labels = json.dumps([s[0] for s in data["sub_statuses"]])
    sub_status_values = json.dumps([s[1] for s in data["sub_statuses"]])

    zone_labels = json.dumps([z[0] for z in data["zones"][:15]])
    zone_values = json.dumps([z[1] for z in data["zones"][:15]])

    partner_labels = json.dumps([p[0] for p in data["partners"][:15]])
    partner_values = json.dumps([p[1] for p in data["partners"][:15]])

    # Aging table rows
    aging_rows = ""
    for label, count in data["buckets"]:
        pct = round(count / data["total"] * 100, 1) if data["total"] else 0
        color_idx = data["all_bucket_labels"].index(label)
        color = (["#3b82f6"] + BUCKET_COLORS)[color_idx]
        aging_rows += f"""
        <tr>
            <td><span class="dot" style="background:{color}"></span>{label}</td>
            <td class="num">{count:,}</td>
            <td class="num">{pct}%</td>
            <td><div class="bar-bg"><div class="bar-fill" style="width:{pct}%;background:{color}"></div></div></td>
        </tr>"""

    # Queue-wise aging breakdown table
    queue_aging_header = "".join(f"<th>{b}</th>" for b in data["all_bucket_labels"])
    queue_aging_rows = ""
    for queue_name in sorted(data["bucket_by_queue"].keys()):
        bq = data["bucket_by_queue"][queue_name]
        total_q = sum(bq.values())
        cells = "".join(f"<td class='num'>{bq.get(b,0)}</td>" for b in data["all_bucket_labels"])
        queue_aging_rows += f"<tr><td class='queue-name'>{queue_name}</td>{cells}<td class='num total'>{total_q}</td></tr>"

    # Top 30 critical tickets table (> 48h, sorted by hours desc)
    critical_tickets = sorted(
        [t for t in data["tickets"] if t["pending_hours"] and t["pending_hours"] > 48],
        key=lambda x: x["pending_hours"],
        reverse=True,
    )[:30]

    critical_rows = ""
    for t in critical_tickets:
        hrs = t["pending_hours"]
        badge_class = "badge-red" if hrs > 120 else "badge-orange" if hrs > 72 else "badge-yellow"
        critical_rows += f"""
        <tr>
            <td>{t['ticket_no']}</td>
            <td>{t['created_date']}</td>
            <td><span class="badge {badge_class}">{hrs:.0f}h</span></td>
            <td>{t['queue']}</td>
            <td>{t['partner']}</td>
            <td>{t['zone'].split(',')[0] if t['zone'] != 'Unknown' else '-'}</td>
            <td>{t['sub_status']}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PFT Internet Issues Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #0f172a; --card: #1e293b; --card2: #334155; --text: #e2e8f0;
    --text2: #94a3b8; --accent: #3b82f6; --green: #22c55e; --red: #ef4444;
    --orange: #f97316; --yellow: #eab308; --border: #475569;
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Segoe UI',system-ui,-apple-system,sans-serif; background:var(--bg); color:var(--text); padding:20px; }}
  .header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:24px; flex-wrap:wrap; gap:12px; }}
  .header h1 {{ font-size:24px; font-weight:700; }}
  .header h1 span {{ color:var(--accent); }}
  .header-right {{ display:flex; gap:12px; align-items:center; }}
  .timestamp {{ color:var(--text2); font-size:13px; }}
  .btn {{ padding:8px 16px; border-radius:8px; text-decoration:none; font-size:13px; font-weight:600;
    border:1px solid var(--border); color:var(--text); background:var(--card); cursor:pointer; transition:all 0.2s; }}
  .btn:hover {{ background:var(--accent); border-color:var(--accent); }}
  .btn-primary {{ background:var(--accent); border-color:var(--accent); color:#fff; }}
  .btn-primary:hover {{ background:#2563eb; }}

  /* Summary Cards */
  .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:16px; margin-bottom:24px; }}
  .card {{ background:var(--card); border-radius:12px; padding:20px; border:1px solid var(--border); }}
  .card-label {{ font-size:12px; color:var(--text2); text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; }}
  .card-value {{ font-size:32px; font-weight:700; }}
  .card-value.green {{ color:var(--green); }}
  .card-value.red {{ color:var(--red); }}
  .card-value.orange {{ color:var(--orange); }}
  .card-value.blue {{ color:var(--accent); }}
  .card-sub {{ font-size:12px; color:var(--text2); margin-top:4px; }}

  /* Charts Grid */
  .charts {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:24px; }}
  .chart-card {{ background:var(--card); border-radius:12px; padding:20px; border:1px solid var(--border); }}
  .chart-card h3 {{ font-size:14px; color:var(--text2); margin-bottom:16px; text-transform:uppercase; letter-spacing:0.5px; }}
  .chart-container {{ position:relative; height:300px; }}

  /* Aging Table */
  .section {{ background:var(--card); border-radius:12px; padding:20px; border:1px solid var(--border); margin-bottom:24px; }}
  .section h3 {{ font-size:14px; color:var(--text2); margin-bottom:16px; text-transform:uppercase; letter-spacing:0.5px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th {{ text-align:left; padding:10px 12px; background:var(--card2); color:var(--text2); font-weight:600;
    font-size:11px; text-transform:uppercase; letter-spacing:0.5px; border-bottom:2px solid var(--border); }}
  td {{ padding:10px 12px; border-bottom:1px solid rgba(71,85,105,0.4); }}
  tr:hover {{ background:rgba(59,130,246,0.05); }}
  .num {{ text-align:right; font-variant-numeric:tabular-nums; font-weight:500; }}
  .total {{ font-weight:700; color:var(--accent); }}
  .queue-name {{ font-weight:600; }}
  .dot {{ display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:8px; }}
  .bar-bg {{ background:var(--card2); border-radius:4px; height:20px; width:100%; overflow:hidden; }}
  .bar-fill {{ height:100%; border-radius:4px; transition:width 0.5s; }}
  .badge {{ padding:3px 8px; border-radius:4px; font-size:11px; font-weight:700; }}
  .badge-red {{ background:rgba(239,68,68,0.2); color:#fca5a5; }}
  .badge-orange {{ background:rgba(249,115,22,0.2); color:#fdba74; }}
  .badge-yellow {{ background:rgba(234,179,8,0.2); color:#fde047; }}

  /* Full width charts */
  .full-width {{ grid-column: 1 / -1; }}

  /* Responsive */
  @media(max-width:900px) {{
    .charts {{ grid-template-columns:1fr; }}
  }}
  @media(max-width:600px) {{
    .cards {{ grid-template-columns:1fr 1fr; }}
    body {{ padding:12px; }}
  }}

  /* Scrollable table */
  .table-scroll {{ overflow-x:auto; }}
</style>
</head>
<body>

<div class="header">
  <h1><span>PFT</span> Internet Issues Dashboard</h1>
  <div class="header-right">
    <span class="timestamp">Report: {data['report_time']}</span>
    <a href="{master_sheet_url}" target="_blank" class="btn btn-primary">Open Master Sheet</a>
    <button class="btn" onclick="window.print()">Export</button>
  </div>
</div>

<!-- Summary Cards -->
<div class="cards">
  <div class="card">
    <div class="card-label">Total Internet Tickets</div>
    <div class="card-value blue">{data['total']:,}</div>
    <div class="card-sub">Filtered from pending report</div>
  </div>
  <div class="card">
    <div class="card-label">Created Today</div>
    <div class="card-value green">{data['uploaded_today']:,}</div>
    <div class="card-sub">New tickets uploaded today</div>
  </div>
  <div class="card">
    <div class="card-label">Critical (&gt; 48h)</div>
    <div class="card-value red">{data['critical']:,}</div>
    <div class="card-sub">{round(data['critical']/data['total']*100,1) if data['total'] else 0}% of total</div>
  </div>
  <div class="card">
    <div class="card-label">In Partner Queue</div>
    <div class="card-value orange">{dict(data['queues']).get('Partner',0):,}</div>
    <div class="card-sub">Waiting on partner action</div>
  </div>
</div>

<!-- Aging Bucket Table -->
<div class="section">
  <h3>Ticket Aging Breakdown (Hours Since Creation)</h3>
  <table>
    <thead><tr><th>Aging Bucket</th><th style="text-align:right">Tickets</th><th style="text-align:right">% Share</th><th>Distribution</th></tr></thead>
    <tbody>{aging_rows}</tbody>
  </table>
</div>

<!-- Charts Row 1 -->
<div class="charts">
  <div class="chart-card">
    <h3>Aging Distribution</h3>
    <div class="chart-container"><canvas id="agingChart"></canvas></div>
  </div>
  <div class="chart-card">
    <h3>Queue-wise Split</h3>
    <div class="chart-container"><canvas id="queueChart"></canvas></div>
  </div>
</div>

<!-- Charts Row 2 -->
<div class="charts">
  <div class="chart-card">
    <h3>Sub Status Breakdown</h3>
    <div class="chart-container"><canvas id="statusChart"></canvas></div>
  </div>
  <div class="chart-card">
    <h3>Top Zones</h3>
    <div class="chart-container"><canvas id="zoneChart"></canvas></div>
  </div>
</div>

<!-- Queue x Aging Cross Table -->
<div class="section">
  <h3>Queue-wise Aging Breakdown</h3>
  <div class="table-scroll">
  <table>
    <thead><tr><th>Queue</th>{queue_aging_header}<th style="text-align:right">Total</th></tr></thead>
    <tbody>{queue_aging_rows}</tbody>
  </table>
  </div>
</div>

<!-- Top Partners -->
<div class="charts">
  <div class="chart-card full-width">
    <h3>Top 15 Partners by Ticket Volume</h3>
    <div class="chart-container" style="height:350px"><canvas id="partnerChart"></canvas></div>
  </div>
</div>

<!-- Critical Tickets Table -->
<div class="section">
  <h3>Top 30 Critical Tickets (&gt; 48 Hours Pending)</h3>
  <div class="table-scroll">
  <table>
    <thead><tr><th>Ticket No</th><th>Created Date</th><th>Pending</th><th>Queue</th><th>Partner</th><th>Zone</th><th>Sub Status</th></tr></thead>
    <tbody>{critical_rows}</tbody>
  </table>
  </div>
</div>

<script>
const chartDefaults = {{
  responsive: true, maintainAspectRatio: false,
  plugins: {{ legend: {{ labels: {{ color:'#94a3b8', font:{{size:11}} }} }} }}
}};

// Aging Bar Chart
new Chart(document.getElementById('agingChart'), {{
  type: 'bar',
  data: {{
    labels: {bucket_labels},
    datasets: [{{ label:'Tickets', data:{bucket_values},
      backgroundColor:{bucket_colors_json}, borderRadius:6, borderSkipped:false }}]
  }},
  options: {{ ...chartDefaults,
    plugins: {{ ...chartDefaults.plugins, legend:{{display:false}} }},
    scales: {{
      x: {{ ticks:{{color:'#94a3b8'}}, grid:{{display:false}} }},
      y: {{ ticks:{{color:'#94a3b8'}}, grid:{{color:'rgba(71,85,105,0.3)'}} }}
    }}
  }}
}});

// Queue Doughnut
new Chart(document.getElementById('queueChart'), {{
  type: 'doughnut',
  data: {{
    labels: {queue_labels},
    datasets: [{{ data:{queue_values},
      backgroundColor:['#3b82f6','#f97316','#22c55e','#eab308','#a855f7','#ec4899'],
      borderWidth:0 }}]
  }},
  options: {{ ...chartDefaults,
    cutout:'60%',
    plugins: {{ ...chartDefaults.plugins,
      legend: {{ position:'right', labels:{{color:'#94a3b8',padding:12,font:{{size:12}}}} }}
    }}
  }}
}});

// Sub Status Doughnut
new Chart(document.getElementById('statusChart'), {{
  type: 'doughnut',
  data: {{
    labels: {sub_status_labels},
    datasets: [{{ data:{sub_status_values},
      backgroundColor:['#3b82f6','#22c55e','#eab308','#ef4444','#a855f7','#ec4899'],
      borderWidth:0 }}]
  }},
  options: {{ ...chartDefaults,
    cutout:'60%',
    plugins: {{ ...chartDefaults.plugins,
      legend: {{ position:'right', labels:{{color:'#94a3b8',padding:12,font:{{size:12}}}} }}
    }}
  }}
}});

// Zone Horizontal Bar
new Chart(document.getElementById('zoneChart'), {{
  type: 'bar',
  data: {{
    labels: {zone_labels},
    datasets: [{{ label:'Tickets', data:{zone_values},
      backgroundColor:'#3b82f6', borderRadius:4, borderSkipped:false }}]
  }},
  options: {{ ...chartDefaults, indexAxis:'y',
    plugins: {{ ...chartDefaults.plugins, legend:{{display:false}} }},
    scales: {{
      x: {{ ticks:{{color:'#94a3b8'}}, grid:{{color:'rgba(71,85,105,0.3)'}} }},
      y: {{ ticks:{{color:'#94a3b8',font:{{size:10}}}}, grid:{{display:false}} }}
    }}
  }}
}});

// Partner Horizontal Bar
new Chart(document.getElementById('partnerChart'), {{
  type: 'bar',
  data: {{
    labels: {partner_labels},
    datasets: [{{ label:'Tickets', data:{partner_values},
      backgroundColor:'#f97316', borderRadius:4, borderSkipped:false }}]
  }},
  options: {{ ...chartDefaults, indexAxis:'y',
    plugins: {{ ...chartDefaults.plugins, legend:{{display:false}} }},
    scales: {{
      x: {{ ticks:{{color:'#94a3b8'}}, grid:{{color:'rgba(71,85,105,0.3)'}} }},
      y: {{ ticks:{{color:'#94a3b8',font:{{size:10}}}}, grid:{{display:false}} }}
    }}
  }}
}});
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Dashboard saved: {output_path}")
    return output_path


def main():
    # Find latest filtered report
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        report_path = sys.argv[1]
    else:
        reports = sorted(glob.glob(os.path.join(SCRIPT_DIR, "internet_issues_tickets_*.xlsx")))
        if not reports:
            print("No filtered Internet Issues report found. Run pft_internet_ticket_agent.py first.")
            sys.exit(1)
        report_path = reports[-1]

    print(f"Loading: {report_path}")

    # Master sheet URL (Kapture health card shared in emails)
    master_sheet_url = "https://docs.google.com/spreadsheets/d/1sijh7r3nVkMWSReKzbp7g9JGoxTR89pPdB0RAbnVh-4/edit"

    data = load_and_analyze(report_path)

    timestamp = datetime.now(IST).strftime("%Y%m%d_%H%M")
    output_path = os.path.join(SCRIPT_DIR, f"dashboard_{timestamp}.html")
    generate_html(data, master_sheet_url, output_path)

    # Also save as latest
    latest_path = os.path.join(SCRIPT_DIR, "dashboard_latest.html")
    generate_html(data, master_sheet_url, latest_path)

    print(f"\nOpen in browser: {latest_path}")
    return latest_path


if __name__ == "__main__":
    main()
