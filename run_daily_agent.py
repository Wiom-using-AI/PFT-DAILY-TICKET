"""
Daily Runner - Automated Email Check + Download + Filter
=========================================================
This script is the daily entry point that:
1. Checks Gmail for today's ~10 AM "Queue Wise Pending Report" email
2. Extracts the download link
3. Downloads the report
4. Filters Internet Issues tickets
5. Saves the output

Can be scheduled via Windows Task Scheduler or run manually.
"""

import subprocess
import json
import re
import os
import sys
import urllib.request
import smtplib
from datetime import datetime, timezone, timedelta

# Add parent dir to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from pft_internet_ticket_agent import (
    download_report,
    filter_internet_tickets,
    log,
    IST,
    DOWNLOAD_DIR,
)

# Gmail API via Google Apps Script or direct IMAP
# For this automation, we use IMAP to fetch the email

import imaplib
import email
from email.header import decode_header


# --- Configuration ---
GMAIL_USER = "avakash.gupta@wiom.in"
# App password (generate at https://myaccount.google.com/apppasswords)
# Store in environment variable for security
GMAIL_APP_PASSWORD_ENV = "GMAIL_APP_PASSWORD"

IMAP_SERVER = "imap.gmail.com"
IMAP_PORT = 993

SENDER_FILTER = "no-reply-report@kapturecrm.com"
# Exact subject line to match — ONLY the first email of the day with this subject
REQUIRED_SUBJECT = "queue wise pending report last 60 days"


def get_app_password():
    """Get Gmail app password from environment variable."""
    password = os.environ.get(GMAIL_APP_PASSWORD_ENV)
    if not password:
        print(f"\nERROR: Gmail App Password not set!")
        print(f"Please set the environment variable: {GMAIL_APP_PASSWORD_ENV}")
        print(f"Steps to generate an App Password:")
        print(f"  1. Go to https://myaccount.google.com/apppasswords")
        print(f"  2. Select 'Mail' and your device")
        print(f"  3. Copy the 16-character password")
        print(f"  4. Set it: set {GMAIL_APP_PASSWORD_ENV}=your_app_password")
        sys.exit(1)
    return password


def search_todays_morning_report():
    """
    Connect to Gmail via IMAP and find today's FIRST email with
    subject 'Queue wise pending report last 60 days'.

    IMPORTANT: Only uses the FIRST email of the day with this exact subject.
    Second/third emails with the same subject are IGNORED.
    """
    password = get_app_password()

    log("Connecting to Gmail via IMAP...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail.login(GMAIL_USER, password)
    mail.select("inbox")

    # Search for today's emails from Kapture with "pending" in subject
    today = datetime.now(IST)
    date_str = today.strftime("%d-%b-%Y")  # e.g., "18-Mar-2026"

    search_criteria = f'(FROM "{SENDER_FILTER}" SINCE "{date_str}" SUBJECT "pending")'
    log(f"IMAP search: {search_criteria}")

    status, message_ids = mail.search(None, search_criteria)
    if status != "OK" or not message_ids[0]:
        log("No matching emails found for today.")
        mail.logout()
        return None

    ids = message_ids[0].split()
    log(f"Found {len(ids)} candidate email(s) today")

    # Collect ALL matching emails with exact subject, then pick the EARLIEST
    matched_emails = []

    for msg_id in ids:
        status, msg_data = mail.fetch(msg_id, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(msg_data[0][1])

        # STRICT subject match: must contain exact phrase
        subject = str(msg.get("Subject", "")).strip()
        subject_lower = subject.lower()
        if REQUIRED_SUBJECT not in subject_lower:
            log(f"  SKIP (wrong subject): '{subject}'")
            continue

        # Parse date
        date_str_raw = msg.get("Date", "")
        try:
            msg_date = email.utils.parsedate_to_datetime(date_str_raw)
            if msg_date.tzinfo is None:
                msg_date = msg_date.replace(tzinfo=timezone.utc)
            msg_date_ist = msg_date.astimezone(IST)
        except Exception:
            continue

        log(f"  MATCH: '{subject}' at {msg_date_ist.strftime('%I:%M %p IST')}")

        # Extract download link from body
        body_text = ""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                if content_type in ("text/plain", "text/html"):
                    payload = part.get_payload(decode=True)
                    if payload:
                        body_text += payload.decode("utf-8", errors="ignore")
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body_text = payload.decode("utf-8", errors="ignore")

        # Extract Kapture download link
        pattern = r'https://storage\.googleapis\.com/kapture_report/EXCEL_Report/[^">\s]+'
        link_match = re.search(pattern, body_text)
        if not link_match:
            log(f"  WARNING: No download link found in this email")
            continue

        matched_emails.append((msg_date_ist, link_match.group(0), subject))

    mail.logout()

    if not matched_emails:
        log("No valid 'Queue wise pending report last 60 days' email found today.")
        return None

    # Sort by time ascending and pick the FIRST (earliest) one
    matched_emails.sort(key=lambda x: x[0])

    first_time, first_link, first_subject = matched_emails[0]
    log(f"")
    log(f"Using FIRST email of the day:")
    log(f"  Subject: '{first_subject}'")
    log(f"  Time: {first_time.strftime('%I:%M %p IST')}")
    log(f"  Link: {first_link}")
    if len(matched_emails) > 1:
        log(f"  NOTE: Ignoring {len(matched_emails) - 1} later email(s) with same subject")
        for i, (t, _, s) in enumerate(matched_emails[1:], 2):
            log(f"    #{i} at {t.strftime('%I:%M %p IST')} — IGNORED")

    return first_link


def main():
    log("=" * 60)
    log("PFT INTERNET TICKET AGENT - DAILY RUN")
    log("=" * 60)

    # Step 1: Find today's email and get download link
    if len(sys.argv) > 1:
        # Allow passing a URL directly for testing
        download_url = sys.argv[1]
        log(f"Using provided URL: {download_url}")
    else:
        download_url = search_todays_morning_report()

    if not download_url:
        log("No email found yet. Will retry every 5 minutes until 12:00 PM IST.")
        return None  # Signal to retry loop

    # Step 2: Download the report
    timestamp = datetime.now(IST).strftime("%Y%m%d")
    filename = f"pending_report_{timestamp}_morning.xlsx"
    report_path = download_report(download_url, filename)

    # Step 3: Filter for Internet Issues
    output_path = filter_internet_tickets(report_path)

    if output_path:
        log("SUCCESS - Internet issues tickets extracted!")
        log(f"Output: {output_path}")

        # Step 4: Save to history database for dashboard
        try:
            from history_db import save_daily_snapshot, save_master_snapshot, get_tickets_for_date
            report_date = datetime.now(IST).strftime("%Y-%m-%d")
            report_time = datetime.now(IST).replace(hour=10, minute=10, second=0, microsecond=0)
            # Count total tickets in full report for total_pending
            import openpyxl as _opx_count
            _wb_count = _opx_count.load_workbook(report_path, read_only=True)
            _ws_count = _wb_count.active
            _total = sum(1 for _ in _ws_count.iter_rows(min_row=2))
            _wb_count.close()
            save_daily_snapshot(output_path, report_date, report_time, total_pending=_total)
            log("History database updated.")
        except Exception as e:
            log(f"Warning: Could not save to history DB: {e}")

        # Step 5: Save category breakdown + full report data for pivot table
        try:
            from history_db import save_category_breakdown, save_full_report, save_queue_category_breakdown
            import openpyxl as _opx
            from collections import defaultdict as _dd
            log("Extracting category breakdown from full report...")
            _wb = _opx.load_workbook(report_path, read_only=True)
            _ws = _wb.active
            _headers = [cell.value for cell in next(_ws.iter_rows(min_row=1, max_row=1))]
            _l3_idx = _headers.index("Disposition Folder Level 3")
            _queue_idx = _headers.index("Current Queue Name")
            _cats = {}
            _queue_bd = _dd(lambda: _dd(int))
            for _row in _ws.iter_rows(min_row=2, values_only=True):
                _val = str(_row[_l3_idx] or "Unknown").strip()
                _cats[_val] = _cats.get(_val, 0) + 1
                _qval = str(_row[_queue_idx] or "(Unknown)").strip()
                _queue_bd[_val][_qval] += 1
            _wb.close()
            save_category_breakdown(report_date, _cats)
            log(f"Category breakdown: {_cats}")
            # Save L3 x Queue breakdown for chart queue filter
            _queue_bd_dict = {l3: dict(queues) for l3, queues in _queue_bd.items()}
            save_queue_category_breakdown(report_date, _queue_bd_dict)
            log(f"Queue breakdown: {len(_queue_bd_dict)} L3 categories with queue splits")

            # Save ALL tickets from full report for category × aging pivot
            log("Saving full report tickets for pivot table...")
            full_count = save_full_report(report_path, report_date, report_time)
            log(f"Full report: {full_count} tickets saved to database")

            # Update total_pending in daily_summary with full report count
            try:
                import sqlite3 as _sql
                _db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ticket_history.db")
                _conn = _sql.connect(_db_path)
                _conn.execute("UPDATE daily_summary SET total_pending = ? WHERE report_date = ? AND (total_pending IS NULL OR total_pending = 0)",
                              (full_count, report_date))
                _conn.commit()
                _conn.close()
                log(f"Updated total_pending = {full_count} for {report_date}")
            except Exception as _e:
                log(f"Warning: Could not update total_pending: {_e}")
        except Exception as e:
            log(f"Warning: Could not save categories/full report: {e}")

        # Step 5c: Save morning resolution snapshot (ticket IDs for resolution tracking)
        try:
            from history_db import save_resolution_snapshot
            log("Saving morning resolution snapshot...")
            _wb2 = _opx.load_workbook(report_path, read_only=True)
            _ws2 = _wb2.active
            _h2 = [cell.value for cell in next(_ws2.iter_rows(min_row=1, max_row=1))]
            _tid_idx = _h2.index("Ticket No")
            _l3_idx2 = _h2.index("Disposition Folder Level 3")
            _morning_ids = []
            _morning_l3 = {}
            for _r in _ws2.iter_rows(min_row=2, values_only=True):
                _tid = str(_r[_tid_idx]).strip() if _r[_tid_idx] else None
                if _tid:
                    _morning_ids.append(_tid)
                    if _r[_l3_idx2]:
                        _morning_l3[_tid] = str(_r[_l3_idx2]).strip()
            _wb2.close()
            save_resolution_snapshot(report_date, "morning", _morning_ids, _morning_l3)
            log(f"Morning resolution snapshot: {len(_morning_ids)} ticket IDs saved")
        except Exception as e:
            log(f"Warning: Could not save morning resolution snapshot: {e}")

        # Step 6: Snapshot master sheet comparison (FIXED at run time)
        try:
            import csv
            import io
            log("Fetching master sheet for comparison snapshot...")
            MASTER_CSV_URL = "https://docs.google.com/spreadsheets/d/1E3Ij57bFHznf3S6cRJSzONaVJ7Tgloud51Z__vXLet0/export?format=csv&gid=1626982265"
            req = urllib.request.Request(MASTER_CSV_URL)
            req.add_header("User-Agent", "Mozilla/5.0")
            response = urllib.request.urlopen(req, timeout=30)
            data = response.read().decode("utf-8-sig")
            reader = csv.reader(io.StringIO(data))
            next(reader)  # skip header
            master_ids = set()
            for row in reader:
                if row and row[0].strip():
                    master_ids.add(row[0].strip())
            log(f"Master sheet has {len(master_ids)} ticket IDs")

            # Get today's ticket IDs from DB
            tickets = get_tickets_for_date(report_date)
            ticket_ids = [t["ticket_no"] for t in tickets]
            result = save_master_snapshot(report_date, master_ids, ticket_ids)
            log(f"Master snapshot: {result['already']} existing, {result['new']} new to upload")

            # Step 6b: Cache new tickets CSV from FULL RAW Excel (all 96 columns)
            try:
                from history_db import save_new_tickets_cache
                new_ids_set = set(result.get("new_ids", []))
                if new_ids_set:
                    log(f"Building full raw CSV for {len(new_ids_set)} new tickets from Excel...")
                    import openpyxl as _opx_raw
                    _wb_raw = _opx_raw.load_workbook(report_path, read_only=True)
                    _ws_raw = _wb_raw.active
                    _raw_headers = [str(cell.value or "") for cell in next(_ws_raw.iter_rows(min_row=1, max_row=1))]
                    _ticket_col = _raw_headers.index("Ticket No") if "Ticket No" in _raw_headers else 0

                    csv_buffer = io.StringIO()
                    writer = csv.writer(csv_buffer)
                    writer.writerow(_raw_headers)

                    new_count = 0
                    for _row in _ws_raw.iter_rows(min_row=2, values_only=True):
                        _tid = str(_row[_ticket_col] or "").strip()
                        if _tid in new_ids_set:
                            writer.writerow([str(v or "") for v in _row])
                            new_count += 1

                    _wb_raw.close()
                    csv_data = csv_buffer.getvalue()
                    save_new_tickets_cache(report_date, csv_data, new_count)
                    log(f"Cached {new_count} new tickets with all {len(_raw_headers)} columns")
                else:
                    log("No new ticket IDs — skipping CSV cache")
            except Exception as ce:
                log(f"Warning: Could not cache new tickets CSV: {ce}")
        except Exception as e:
            log(f"Warning: Could not snapshot master sheet: {e}")

        # Step 7: Cleanup old data (tickets: 30 days, summaries: 90 days)
        try:
            from history_db import cleanup_old_data
            cleanup_old_data()
        except Exception as e:
            log(f"Warning: Could not cleanup old data: {e}")
    else:
        log("No Internet Issues tickets found in today's report.")

    return output_path


def run_with_retry():
    """Run the daily agent with retry logic.
    If the email hasn't arrived, retry every 5 minutes until 12:00 PM IST.
    """
    import time

    RETRY_INTERVAL = 5 * 60  # 5 minutes in seconds
    DEADLINE_HOUR = 12  # 12:00 PM IST

    attempt = 1
    while True:
        now = datetime.now(IST)
        log(f"Attempt #{attempt} at {now.strftime('%I:%M %p IST')}")

        result = main()

        if result is not None:
            # Success — email found and processed
            log("Daily agent completed successfully.")
            return result

        # Check if we've passed the deadline
        if now.hour >= DEADLINE_HOUR:
            log(f"DEADLINE REACHED ({DEADLINE_HOUR}:00 PM IST). Email not found today.")
            log("You may need to manually download and process the report.")
            sys.exit(1)

        # Wait and retry
        next_try = now + timedelta(seconds=RETRY_INTERVAL)
        log(f"Retrying at {next_try.strftime('%I:%M %p IST')}...")
        time.sleep(RETRY_INTERVAL)
        attempt += 1


if __name__ == "__main__":
    run_with_retry()
