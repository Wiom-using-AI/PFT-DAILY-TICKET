"""
PFT Agent - Internet Ticket Filter Automation
==============================================
This script:
1. Searches Gmail for the daily "Queue Wise Pending Report" email (~10 AM IST)
2. Extracts the download link from the email body
3. Downloads the attached .xlsx report
4. Filters tickets where "Disposition Folder Level 3" == "Internet Issues"
5. Saves the filtered tickets to a new Excel file with timestamp
"""

import subprocess
import json
import re
import os
import sys
import urllib.request
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Configuration ---
DOWNLOAD_DIR = os.path.dirname(os.path.abspath(__file__))
IST = timezone(timedelta(hours=5, minutes=30))
FILTER_COLUMN = "Disposition Folder Level 3"
FILTER_VALUE = "Internet Issues"
GMAIL_SENDER = "no-reply-report@kapturecrm.com"
SUBJECT_KEYWORD = "pending"


def log(msg):
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S IST")
    print(f"[{timestamp}] {msg}")


def search_gmail_for_report():
    """Search Gmail for today's morning pending report."""
    today_str = datetime.now(IST).strftime("%Y/%m/%d")
    query = f"from:{GMAIL_SENDER} subject:{SUBJECT_KEYWORD} after:{today_str}"

    log(f"Searching Gmail with query: {query}")

    # Use Gmail search via MCP - we'll call the claude CLI or use the API
    # For standalone execution, we search and find the download link
    # Try using the gmail API through the MCP tool via subprocess
    # Fallback: search for today's emails matching the pattern

    return query


def extract_download_link(email_body):
    """Extract the Kapture report download URL from email body."""
    # Pattern: href="https://storage.googleapis.com/kapture_report/EXCEL_Report/..."
    pattern = r'https://storage\.googleapis\.com/kapture_report/EXCEL_Report/[^">\s]+'
    match = re.search(pattern, email_body)
    if match:
        return match.group(0)
    return None


def download_report(url, filename=None):
    """Download the xlsx report from the given URL."""
    if filename is None:
        timestamp = datetime.now(IST).strftime("%Y%m%d_%H%M")
        filename = f"pending_report_{timestamp}.xlsx"

    filepath = os.path.join(DOWNLOAD_DIR, filename)
    log(f"Downloading report to: {filepath}")

    urllib.request.urlretrieve(url, filepath)

    file_size = os.path.getsize(filepath) / (1024 * 1024)
    log(f"Downloaded successfully ({file_size:.1f} MB)")
    return filepath


def filter_internet_tickets(input_path):
    """Filter tickets where Disposition Folder Level 3 == 'Internet Issues'."""
    log(f"Reading report: {input_path}")
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb.active

    # Find headers
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    # Find the filter column index
    try:
        filter_col_idx = headers.index(FILTER_COLUMN)
    except ValueError:
        log(f"ERROR: Column '{FILTER_COLUMN}' not found in the report!")
        log(f"Available columns: {[h for h in headers if h]}")
        wb.close()
        return None

    log(f"Found '{FILTER_COLUMN}' at column index {filter_col_idx}")

    # Collect filtered rows
    filtered_rows = []
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        cell_value = row[filter_col_idx]
        if cell_value and str(cell_value).strip() == FILTER_VALUE:
            filtered_rows.append(row)

    wb.close()
    log(f"Total tickets in report: {total_rows}")
    log(f"Internet Issues tickets found: {len(filtered_rows)}")

    if not filtered_rows:
        log("No Internet Issues tickets found.")
        return None

    # Create output workbook with formatting
    timestamp = datetime.now(IST).strftime("%Y%m%d_%H%M")
    output_filename = f"internet_issues_tickets_{timestamp}.xlsx"
    output_path = os.path.join(DOWNLOAD_DIR, output_filename)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Internet Issues"

    # Style definitions
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write filtered data
    for row_idx, row_data in enumerate(filtered_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = out_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-fit key columns (approximate widths)
    key_columns = {
        "Ticket No": 18,
        "Created Date": 14,
        "Customer Name": 25,
        "Disposition Folder Level 3": 25,
        "Status": 15,
        "Channel Partner": 25,
        "Mapped Partner name": 25,
        "City": 15,
        "Zone": 15,
        "Pending No of Days": 18,
        "Current Queue Name": 22,
        "Device ID": 20,
    }
    for col_idx, header in enumerate(headers, 1):
        if header in key_columns:
            out_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = key_columns[header]

    # Freeze top row
    out_ws.freeze_panes = "A2"

    # Add auto-filter
    out_ws.auto_filter.ref = out_ws.dimensions

    out_wb.save(output_path)
    log(f"Filtered report saved: {output_path}")

    # Print summary
    print("\n" + "=" * 60)
    print(f"  INTERNET ISSUES TICKET SUMMARY")
    print(f"  Date: {datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}")
    print(f"  Total Pending Tickets: {total_rows}")
    print(f"  Internet Issues Tickets: {len(filtered_rows)}")
    print(f"  Output File: {output_filename}")
    print("=" * 60 + "\n")

    return output_path


def run_with_download_link(download_url):
    """Full pipeline: download report and filter internet tickets."""
    log("=== PFT Internet Ticket Agent Started ===")

    # Step 1: Download
    report_path = download_report(download_url)

    # Step 2: Filter
    output_path = filter_internet_tickets(report_path)

    if output_path:
        log("=== Agent completed successfully ===")
    else:
        log("=== Agent completed - no internet tickets found ===")

    return output_path


def run_on_local_file(filepath):
    """Run filter on an already-downloaded report file."""
    log("=== PFT Internet Ticket Agent - Local File Mode ===")
    output_path = filter_internet_tickets(filepath)
    if output_path:
        log("=== Agent completed successfully ===")
    else:
        log("=== Agent completed - no internet tickets found ===")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.startswith("http"):
            run_with_download_link(arg)
        elif os.path.isfile(arg):
            run_on_local_file(arg)
        else:
            print(f"Usage: python {sys.argv[0]} <download_url_or_local_file>")
            print(f"  Provide either a Kapture report download URL or local .xlsx file path")
    else:
        # Default: run on any pending_report*.xlsx in current directory
        import glob
        reports = sorted(glob.glob(os.path.join(DOWNLOAD_DIR, "pending_report*.xlsx")))
        if reports:
            run_on_local_file(reports[-1])
        else:
            print("No report file found. Provide a download URL or local file path.")
            print(f"Usage: python {sys.argv[0]} <download_url_or_local_file>")
